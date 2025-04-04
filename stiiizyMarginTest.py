import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

stiiizy_deal = {
    'vendors': ['Elevation (Stiiizy)'],
    'brands': ['Stiiizy'],
    'discount': 0.40,
    'kickback': 0.30,
    'weekend_days': ['Thursday', 'Friday', 'Saturday'],
    'weekend_kickback_categories': ["Eighths", "Pre-Rolls", "Flower", "Halves", "Quarters", "Ounces", "Concentrate", "Accessories"],
    'always_kickback_categories': ["Disposables", "Cartridges", "Gummies", "Edibles"]
}

stiiizy_deal = {
    'vendors': ['Varavo'],
    'brands': ['Kushy Punch'],
    'discount': 0.30,
    'kickback': 0.0,
    'weekend_days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
    'weekend_kickback_categories': ["Disposables", "Gummies", "Edibles"],
    'always_kickback_categories': ["Disposables", "Gummies", "Edibles"]
}

def ensure_dir_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

def format_excel_file(filename: str):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
    wb.save(filename)

def calculate_margin(cost, price, kickback_applicable):
    revenue = price * (1 - stiiizy_deal['discount'])
    kickback = cost * stiiizy_deal['kickback'] if kickback_applicable else 0
    net_cost = cost - kickback
    if revenue == 0:
        return 0
    return round(((revenue - net_cost) / revenue) * 100, 2)

def target_price_for_45_margin(cost, kickback_applicable):
    kickback = cost * stiiizy_deal['kickback'] if kickback_applicable else 0
    net_cost = cost - kickback
    revenue_needed = net_cost / (1 - 0.45)
    price_needed = revenue_needed / (1 - stiiizy_deal['discount'])
    return round(price_needed, 2)

def process_stiiizy_margins():
    input_path = os.path.join("files", "04-03-2025_MV.csv")
    sales_data_path = os.path.join("files", "salesMV.xlsx")
    output_directory = "done"
    ensure_dir_exists(output_directory)

    try:
        df = pd.read_csv(input_path)
    except Exception as e:
        print(f"Error reading file: {e}")
        return

    df = df[df['Brand'].isin(stiiizy_deal['brands']) & df['Vendor'].isin(stiiizy_deal['vendors'])]
    if df.empty:
        print("No matching Stiiizy products found.")
        return

    try:
        sales_df = pd.read_excel(sales_data_path, header=4)
        sales_df.columns = sales_df.columns.str.strip().str.lower()
        sales_df.rename(columns={'product name': 'Product', 'total inventory sold': 'Units Sold'}, inplace=True)
        sales_df['order time'] = pd.to_datetime(sales_df['order time'], errors='coerce')
        sales_df['day of week'] = sales_df['order time'].dt.strftime('%A')
    except Exception as e:
        print(f"Warning: Could not load sales data: {e}")
        sales_df = pd.DataFrame()

    df = df.merge(sales_df.groupby('Product', as_index=False)['Units Sold'].sum(), on='Product', how='left')
    df['Units Sold'] = df['Units Sold'].fillna(0)
    df = df[df['Units Sold'] > 0]

    def get_margins(row):
        category = row['Category']
        cost = row['Cost']
        price = row['Price']
        product = row['Product']

        wknd = category in stiiizy_deal['always_kickback_categories'] or category in stiiizy_deal['weekend_kickback_categories']
        wkdy = category in stiiizy_deal['always_kickback_categories']

        margin_wknd = calculate_margin(cost, price, wknd)
        margin_wkdy = calculate_margin(cost, price, wkdy)

        # Calculate average margin from actual sales days
        matching_sales = sales_df[sales_df['Product'] == product]
        total_units = 0
        weighted_margin = 0

        for _, sale in matching_sales.iterrows():
            day = sale['day of week']
            units = sale['Units Sold']
            kickback = category in stiiizy_deal['always_kickback_categories'] or (day in stiiizy_deal['weekend_days'] and category in stiiizy_deal['weekend_kickback_categories'])
            margin = calculate_margin(cost, price, kickback)
            weighted_margin += margin * units
            total_units += units

        avg_margin = weighted_margin / total_units if total_units > 0 else 0

        # Calculate target price using avg_margin
        margin_rate = avg_margin / 100
        revenue_needed = cost / (1 - stiiizy_deal['kickback']) / (1 - margin_rate) if margin_rate < 1 else 0
        target_price = revenue_needed / (1 - stiiizy_deal['discount']) if revenue_needed else 0

        return pd.Series([margin_wknd, margin_wkdy, round(avg_margin, 2), round(target_price, 2)])

    df[['Margin_Thu-Sat_%', 'Margin_Sun-Wed_%', 'Avg Margin From Sales Days (%)', 'TargetPrice_45Margin']] = df.apply(get_margins, axis=1)

    columns = ['Product', 'Category', 'Cost', 'Price', 'Margin_Thu-Sat_%', 'Margin_Sun-Wed_%', 'Avg Margin From Sales Days (%)', 'TargetPrice_45Margin', 'Units Sold']
    df = df[columns]
    df.sort_values(by=['Price', 'Product'], inplace=True)

    today_str = datetime.now().strftime("%m-%d-%Y")
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    out_dir = os.path.join(output_directory, base_name)
    ensure_dir_exists(out_dir)

    output_file = os.path.join(out_dir, f"Stiiizy_Margins_{today_str}.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Stiiizy Margins')
    format_excel_file(output_file)
    print(f"Saved report: {output_file}")

if __name__ == "__main__":
    process_stiiizy_margins()
