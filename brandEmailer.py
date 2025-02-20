#!/usr/bin/env python3

import os
import openpyxl
import re

# ---------------------------------------------
# 1) GMAIL API SEND LOGIC (From your autoJob.py)
# ---------------------------------------------
def send_email_with_gmail_html(subject, html_body, recipients, attachments=None):
    """
    Sends an HTML email via Gmail API with optional attachments.
    Adjust SCOPES, token filenames, etc. to match your environment.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    creds = None
    gmail_token = "token_gmail.json"  # or whichever token file you normally use

    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token, "w") as f:
            f.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    # Create a MIMEMultipart message for HTML
    msg = MIMEMultipart('alternative')
    msg['From'] = "me"  # The Gmail API ignores this 'From', but it's good practice
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    # Attach the HTML body
    part_html = MIMEText(html_body, 'html')
    msg.attach(part_html)

    # Optionally attach files
    if attachments:
        for file_path in attachments:
            if not os.path.isfile(file_path):
                continue
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as fp:
                file_data = fp.read()
            part = MIMEBase("application", "octet-stream")
            part.set_payload(file_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_req = {'raw': raw_message}

    try:
        sent = service.users().messages().send(userId='me', body=send_req).execute()
        print(f"[INFO] HTML Email sent! ID: {sent['id']} | Subject: {subject}")
    except Exception as e:
        print(f"[ERROR] Could not send HTML email: {e}")

# ---------------------------------------------
# 2) BRAND -> EMAIL MAP
# ---------------------------------------------
# In production, you might store real brand-specific emails. 
# For now, default everything to "anthony@barbaro.tech" for testing.

BRAND_EMAILS = {
    # "Hashish": "team@hashish.com",
    # "Jeeter": "sales@jeeterbrand.com",
    # ...
    # Provide a default fallback if brand not in dictionary:
}

DEFAULT_EMAIL = "anthony@barbaro.tech"

# ---------------------------------------------
# 3) HELPER: PARSE "SUMMARY" SHEET FOR (STORE, KICKBACK OWED)
# ---------------------------------------------
def parse_kickback_summary(brand_report_path):
    """
    Looks for a 'Summary' sheet in the given xlsx. 
    Returns a list of (store, owed) floats or strings.
    """
    results = []
    if not os.path.isfile(brand_report_path):
        return results

    wb = openpyxl.load_workbook(brand_report_path, data_only=True)
    if "Summary" not in wb.sheetnames:
        wb.close()
        return results

    sh = wb["Summary"]
    # By your deals.py code, column A => "Store", B => "Kickback Owed"
    # Data starts at row 2 onward
    for row_idx in range(2, sh.max_row + 1):
        store_val = sh.cell(row=row_idx, column=1).value
        owed_val  = sh.cell(row=row_idx, column=2).value

        if store_val is not None and owed_val is not None:
            # Filter out obvious header or empty
            # "Store" in store_val => skip, "Kickback Owed" in owed_val => skip
            str_store = str(store_val).strip()
            str_owed  = str(owed_val).strip().lower()
            if (str_store.lower() not in ["store", ""] and 
                str_owed not in ["kickback owed", ""]):
                results.append((store_val, owed_val))
    wb.close()
    return results

def build_kickback_table(rows):
    """
    Convert list[(store, owed_value), ...] into a small HTML table.
    """
    if not rows:
        return "<p>(No data found in summary.)</p>"

    html = "<table border='1' cellpadding='5' cellspacing='0'>"
    html += "<thead><tr><th>Store</th><th>Kickback Owed</th></tr></thead><tbody>"
    for (store, owed) in rows:
        # Format numeric if possible
        try:
            owed_float = float(owed)
            owed_str = f"${owed_float:,.2f}"
        except:
            owed_str = str(owed)
        html += f"<tr><td>{store}</td><td>{owed_str}</td></tr>"
    html += "</tbody></table>"
    return html

# ---------------------------------------------
# 4) OPTIONAL: PARSE links.txt FOR BRAND LINKS
# ---------------------------------------------
def load_links_map(links_file="links.txt"):
    """
    Reads each line of links.txt: 'filename.xlsx: https://drive.google...'
    Returns a dict:
        {
          "Hashish": [ (full_line_1), (full_line_2), ...],
          "Jeeter": [...],
          ...
        }

    We'll match brand name if the filename starts with that brand prefix, 
    e.g. "Hashish_report..." => brand=Hashish
    """
    brand_links_map = {}
    if not os.path.isfile(links_file):
        return brand_links_map

    with open(links_file, "r", encoding="utf-8") as f:
        lines = [ln.strip() for ln in f if ln.strip()]

    for line in lines:
        # line = "Hashish_report_2025-01-01_to_2025-01-07.xlsx: https://drive.google.com/..."
        # brand is typically the first chunk up to "_report" ...
        # We'll do a quick approach:
        match = re.match(r"^([^_\s]+)_report_.*:\s*(https?://\S+)$", line, re.IGNORECASE)
        if match:
            brand = match.group(1)
            # Some brands might have slashes or spaces replaced. If your brand has tricky characters,
            # you can do a more robust parse. For now, we assume brand is "Hashish", "Jeeter", etc.
            brand = brand.replace("-", "").replace(" ", "")

            if brand not in brand_links_map:
                brand_links_map[brand] = []
            brand_links_map[brand].append(line)
        else:
            # Could not parse brand. Possibly the consolidated file or a different naming pattern.
            pass

    return brand_links_map

def make_html_link_list(lines):
    """
    Turn lines like: 
      "Hashish_report_2025-01-01_to_2025-01-07.xlsx: https://drive..."
    into HTML <ul> items with clickable link
    """
    if not lines:
        return "<p>No links for this brand.</p>"

    html = "<ul>"
    for ln in lines:
        if ":" in ln:
            filename, link = ln.split(":", 1)
            filename = filename.strip()
            link = link.strip()
            html += f"<li><strong>{filename}</strong>: <a href='{link}'>{link}</a></li>"
        else:
            html += f"<li>{ln}</li>"
    html += "</ul>"
    return html

# ---------------------------------------------
# 5) MAIN “SEND BRAND EMAILS” LOGIC
# ---------------------------------------------
def send_brand_emails():
    """
    Looks in 'brand_reports/' for each brand .xlsx file (e.g. "Hashish_report_2025-xx-xx.xlsx").
    For each brand file:
      - Parse brand name from filename
      - Gather Kickback Owed data from the "Summary" sheet
      - (Optional) gather brand's GDrive link(s) from links.txt
      - Send an HTML email to the brand's contact (or default) with:
          * Subject: "Weekly Kickback - [Brand]"
          * Body: Kickback table + links
          * Attachment: the brand .xlsx (optional)
    Finally, optionally send a single "consolidated" email with all brand data.
    """
    reports_dir = "brand_reports"
    if not os.path.isdir(reports_dir):
        print(f"[ERROR] The folder '{reports_dir}' does not exist. No emails sent.")
        return
    
    # Load brand->[lines] from links.txt if you want clickable GDrive links
    brand_links_map = load_links_map("links.txt")

    # We'll keep track of all brand summaries so we can do a consolidated email if you want
    all_brands_info = []

    for filename in os.listdir(reports_dir):
        if not filename.endswith(".xlsx"):
            continue
        # By deals.py naming, it should be "Brand_report_YYYY-mm-dd_to_YYYY-mm-dd.xlsx"
        # Let's parse brand name up to "_report"
        match = re.match(r"^([^_]+)_report_.*\.xlsx$", filename, re.IGNORECASE)
        if not match:
            continue
        brand_name = match.group(1)

        # The path to attach
        file_path = os.path.join(reports_dir, filename)

        # Parse (store, owed) from summary
        summary_rows = parse_kickback_summary(file_path)
        owed_table_html = build_kickback_table(summary_rows)

        # Optionally get brand's GDrive link lines from brand_links_map
        # Normalize brand name if needed (like 'Smackers' vs 'Smackers' key).
        # For safety, compare case-insensitive:
        normalized_brand = brand_name.lower().replace("-", "").replace(" ", "")

        brand_link_lines = brand_links_map.get(normalized_brand, [])
        link_list_html = make_html_link_list(brand_link_lines)

        # Decide the recipient
        recipient = BRAND_EMAILS.get(brand_name, DEFAULT_EMAIL)

        # Build HTML body
        html_body = f"""
        <html>
          <body>
            <p>Hello {brand_name},</p>
            <p>Please see below your Kickback Owed details and the Drive link(s):</p>
            <h3>Kickback Summary:</h3>
            {owed_table_html}
            <h3>Google Drive Link(s):</h3>
            {link_list_html}
            <p>Attached is your brand sales report as well, if needed.</p>
            <p>Regards,<br>Anthony</p>
          </body>
        </html>
        """

        subject = f"Weekly Kickback - {brand_name}"
        print(f"[INFO] Sending email for brand '{brand_name}' to '{recipient}' ...")
        send_email_with_gmail_html(
            subject=subject,
            html_body=html_body,
            recipients=recipient,
            attachments=[file_path]  # or [] if you don't want to attach
        )

        # Collect info for consolidated
        total_owed = 0.0
        for (_, o) in summary_rows:
            try:
                total_owed += float(o)
            except:
                pass
        all_brands_info.append({
            "brand": brand_name,
            "rows": summary_rows,
            "links": brand_link_lines,
            "total_owed": total_owed
        })

    # ---------------------------------------------
    # 6) OPTIONAL: SEND CONSOLIDATED EMAIL
    # ---------------------------------------------
    # If you want a single email that shows all brands in one table,
    # plus all brand links. Just for your internal use perhaps.
    if all_brands_info:
        grand_total = sum(bi["total_owed"] for bi in all_brands_info)
        # Build HTML for each brand
        brand_htmls = []
        for bi in all_brands_info:
            brand = bi["brand"]
            link_html = make_html_link_list(bi["links"])
            row_html = build_kickback_table(bi["rows"])
            brand_html = f"""
            <h3>{brand}</h3>
            <p><strong>Links:</strong></p>
            {link_html}
            <p><strong>Kickback Summary:</strong></p>
            {row_html}
            """
            brand_htmls.append(brand_html)

        consolidated_html = "<hr>".join(brand_htmls)
        consolidated_html += f"<p><strong>GRAND TOTAL OWED: ${grand_total:,.2f}</strong></p>"

        subject = "Consolidated Weekly Kickback (All Brands)"
        html_body = f"""
        <html>
        <body>
          <p>Hello Team,</p>
          <p>Below is a consolidated summary of all brands Kickback info:</p>
          {consolidated_html}
          <p>Regards,<br>Anthony</p>
        </body>
        </html>
        """

        # Send to yourself or whomever
        send_email_with_gmail_html(
            subject=subject,
            html_body=html_body,
            recipients=["anthony@barbaro.tech"],  # or a distribution list
            attachments=None
        )

# ---------------------------------------------
# 7) MAIN ENTRY POINT
# ---------------------------------------------
if __name__ == "__main__":
    send_brand_emails()
