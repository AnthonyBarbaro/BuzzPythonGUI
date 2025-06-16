#!/usr/bin/env python3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ——— Configuration ———
INPUT_FILE        = "files/Discount-Detail-Report-12_1_2024-6_14_2025.xlsx"
OUTPUT_FILE       = "weedmaps_report.xlsx"
DESC_FILTER       = "Weedmaps 2% online order"
PERCENT_THRESHOLD = 2.0

KEEP_COLS = [
    "Location Name",
    "Order ID",
    "Order Time",
    "Customer Name",
    "Product Name",
    "Unit Price",
    "Gross Sales",
    "Discounted Amount",
    "Discount Percent",
    "Net Sales",
    "Discount Name",
    "Budtender Name",
    "Discount Approved By",
]
# ————————————

def main():
    # 1. Load using row 5 as header (zero-based index 4)
    df = pd.read_excel(INPUT_FILE, header=4, engine="openpyxl")

    # 2. Normalize & convert Discount Percent to float
    df["Discount Percent"] = (
        df["Discount Percent"]
          .astype(str)
          .str.rstrip("%")
          .replace("", "0")
          .astype(float)
    )

    # 3. Filter by description & threshold
    mask_desc = df["Discount Description"].astype(str)\
                  .str.contains(DESC_FILTER, case=False, na=False)
    mask_pct  = df["Discount Percent"] > PERCENT_THRESHOLD
    filtered  = df[mask_desc & mask_pct].copy()

    # 4. Sort descending
    filtered.sort_values("Discount Percent", ascending=False, inplace=True)

    # 5. Compute Usage Count per customer
    filtered["Usage Count"] = (
        filtered.groupby("Customer Name")["Customer Name"]
                .transform("count")
    )

    # 6. Prepare detail sheet
    detail_cols = KEEP_COLS + ["Usage Count"]
    detail_df   = filtered[detail_cols]

    # 7. Build approver summary
    approver_summary = (
        filtered
        .groupby("Discount Approved By")
        .agg(
            Total_Approvals=("Discount Approved By", "size"),
            Distinct_Times  =("Order Time", "nunique")
        )
        .reset_index()
    )
    abuse_df = approver_summary[
        approver_summary["Total_Approvals"] > approver_summary["Distinct_Times"]
    ]

    # 8. Write to Excel with abuse sheet first
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        abuse_df.to_excel(writer, sheet_name="Approver Abuse", index=False)
        detail_df.to_excel(writer, sheet_name="All Weedmaps >2%", index=False)

    # 9. Auto‐fit columns & freeze headers
    wb = load_workbook(OUTPUT_FILE)
    for sheet in wb.worksheets:
        max_row = sheet.max_row
        max_col = sheet.max_column

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, max_row + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    length = len(str(val))
                    if length > max_length:
                        max_length = length
            sheet.column_dimensions[col_letter].width = max_length + 2

        sheet.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"✔ Report written to '{OUTPUT_FILE}' with 'Approver Abuse' first.")

if __name__ == "__main__":
    main()