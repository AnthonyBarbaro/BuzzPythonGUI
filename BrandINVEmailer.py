#!/usr/bin/env python3
"""
BrandINVEmailer.py

Combined script that:

1. Reads JSON config (including test_mode, test_email, brand list).
2. On the current day, fetches CSVs (via getCatalog.py) [optional].
3. Processes brand inventory, generating Excel files grouped by brand.
4. Uploads those files to a date-based folder inside a parent "INVENTORY" Google Drive folder
   - Each brand gets its own subfolder (made public).
5. Sends an HTML email to the brand's recipients with a single publicly shareable
   Drive folder link (since the folder is public, all files are accessible).

Requires:
- credentials.json (for Google OAuth Drive + Gmail)
- brand_config.json (for daily brand scheduling, plus test-mode toggle)
"""

import os
import sys
import json
import subprocess
import datetime
import traceback
import shutil
import re
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ------------------------------------------------------------------------------
# ------------------------- CONFIG / CONSTANTS ----------------------------------
# ------------------------------------------------------------------------------

# Folders for CSV input and XLSX output
INPUT_DIRECTORY = "files"       # Where CSVs land
LOCAL_REPORTS_FOLDER = "brand_reports_tmp"  # Local subfolder for generated reports

BRAND_CONFIG_JSON = "brand_config.json"

# Google Drive parent folder name (where we create subfolders by date)
DRIVE_PARENT_FOLDER_NAME = "INVENTORY"

# OAuth credential files
CREDENTIALS_FILE = "credentials.json"
TOKEN_DRIVE_FILE = "token_drive.json"   # Stores Drive API tokens
TOKEN_GMAIL_FILE = "token_gmail.json"   # Stores Gmail API tokens

# Google Drive API Scopes
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Gmail API Scopes
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

# ------------------------------------------------------------------------------
# --------------------- GMAIL API SEND HTML HELPER -----------------------------
# ------------------------------------------------------------------------------

def gmail_authenticate():
    """
    Authenticate with Gmail API using OAUTH and return a service object.
    """
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_GMAIL_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_GMAIL_FILE, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_GMAIL_FILE, "w") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def send_email_with_gmail_html(subject, html_body, recipients):
    """
    Sends an HTML email via the Gmail API. 
    `recipients` can be a list or a single string.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if isinstance(recipients, str):
        recipients = [recipients]

    service = gmail_authenticate()

    msg = MIMEMultipart("alternative")
    msg["From"] = "me"  # 'me' means authenticated user
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    part_html = MIMEText(html_body, "html")
    msg.attach(part_html)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    body = {"raw": raw_message}

    try:
        sent = service.users().messages().send(userId="me", body=body).execute()
        print(f"[GMAIL] Email sent! ID: {sent['id']} | Subject: {subject}")
    except Exception as e:
        print(f"[ERROR] Could not send HTML email via Gmail API: {e}")


# ------------------------------------------------------------------------------
# ------------------------- GOOGLE DRIVE HELPER ---------------------------------
# ------------------------------------------------------------------------------

def drive_authenticate():
    """
    Authenticate & build the Google Drive service using OAuth.
    """
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_DRIVE_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_DRIVE_FILE, DRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_DRIVE_FILE, "w") as token:
            token.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


def make_folder_public(service, folder_id):
    """
    Make the given Google Drive folder public (viewable by anyone with the link).
    """
    try:
        permission = {
            "type": "anyone",
            "role": "reader"
        }
        service.permissions().create(fileId=folder_id, body=permission).execute()
        print(f"[INFO] Folder ID {folder_id} is now public.")
    except Exception as e:
        print(f"[ERROR] Could not make folder public: {e}")


def find_or_create_folder(service, folder_name, parent_id=None):
    """
    Find a folder named `folder_name` under `parent_id` (if given).
    If not found, create it.
    Returns the folder's ID.
    Only makes the folder public if it was newly created.
    """
    from googleapiclient.errors import HttpError
    
    folder_name_escaped = folder_name.replace("'", "\\'")
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name_escaped}'"
    if parent_id:
        query += f" and '{parent_id}' in parents"

    try:
        response = service.files().list(q=query, spaces="drive", fields="files(id, name)").execute()
        folders = response.get("files", [])
    except HttpError as err:
        print(f"[ERROR] Drive folder lookup failed: {err}")
        return None

    # If folder exists, return its ID directly (do NOT make it public again)
    if folders:
        folder_id = folders[0]["id"]
        return folder_id

    # Otherwise, create the folder
    folder_metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder"
    }
    if parent_id:
        folder_metadata["parents"] = [parent_id]

    new_folder = service.files().create(body=folder_metadata, fields="id").execute()
    folder_id = new_folder.get("id")
    print(f"[INFO] Created new folder '{folder_name}' (ID: {folder_id})")

    # Now make it public exactly once (since it's newly created)
    try:
        make_folder_public(service, folder_id)
    except Exception as e:
        print(f"[ERROR] Unable to set public permission on new folder: {e}")

    return folder_id


def upload_file_to_drive(service, file_path, folder_id):
    """
    Upload local file `file_path` to Google Drive in `folder_id`. Return file ID.
    """
    from googleapiclient.http import MediaFileUpload

    file_name = os.path.basename(file_path)
    file_metadata = {
        "name": file_name,
        "parents": [folder_id]
    }
    media = MediaFileUpload(file_path, resumable=True)
    drive_file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id"
    ).execute()
    return drive_file.get("id")


# ------------------------------------------------------------------------------
# ---------------------- INVENTORY PROCESSING FUNCTIONS -------------------------
# ------------------------------------------------------------------------------

INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand']

def safe_makedirs(path):
    """Create directory if it doesn't exist."""
    if not os.path.exists(path):
        os.makedirs(path)

def extract_strain_type(product_name: str):
    if not isinstance(product_name, str):
        return ""
    name = " " + product_name.upper() + " "
    if re.search(r'\bS\b', name):
        return 'S'
    if re.search(r'\bH\b', name):
        return 'H'
    if re.search(r'\bI\b', name):
        return 'I'
    return ""

def extract_product_details(product_name: str):
    if not isinstance(product_name, str):
        return "", ""
    name_upper = product_name.upper()

    weight_match = re.search(r'(\d+(\.\d+)?)G', name_upper)
    weight = weight_match.group(0) if weight_match else ""
    sub_type = ""
    if " HH " in f" {name_upper} ":
        sub_type = "HH"
    elif " IN " in f" {name_upper} ":
        sub_type = "IN"
    return weight, sub_type

def is_empty_or_numbers(val):
    if not isinstance(val, str):
        return True
    val_str = val.strip()
    return val_str == "" or val_str.isdigit()

def format_excel_file(filename: str):
    """
    **ADVANCED** Excel formatting:  
    1) Freeze header row,  
    2) Bold + fill header,  
    3) Auto-fit columns,  
    4) Insert category rows for 'Category' changes,  
    5) Etc.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(filename)

    for ws in wb.worksheets:
        # Freeze the first row
        ws.freeze_panes = "A2"

        # Format the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Optional: find columns by name
        available_col = None
        category_col = None
        strain_col = None
        product_col = None
        brand_col = None
        for i, cell in enumerate(ws[1], start=1):
            val = (cell.value or '').lower()
            if val == 'category':
                category_col = i
            elif val == 'available':
                available_col = i
            elif val == 'product':
                product_col = i
            elif val == 'brand':
                brand_col = i
            elif val == 'strain_type':
                strain_col = i

        # Widen the "Available" column
        if available_col:
            col_letter = get_column_letter(available_col)
            if ws.column_dimensions[col_letter].width < 20:
                ws.column_dimensions[col_letter].width = 20

        # Insert grouping rows whenever the Category changes
        if category_col:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            current_type = None
            insert_positions = []
            for idx, row_data in enumerate(rows, start=2):
                cat_val = row_data[category_col - 1]
                if cat_val != current_type:
                    if current_type is not None:
                        insert_positions.append(idx)
                    current_type = cat_val
            if rows:
                # Insert a row at top if there's data
                insert_positions.insert(0, 2)

            category_font = Font(bold=True, size=14)
            fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            current_type = None
            group_types = []
            row_counter = 2
            for row_data in rows:
                cat_val = row_data[category_col - 1]
                if cat_val != current_type:
                    group_types.append((row_counter, cat_val))
                    current_type = cat_val
                row_counter += 1

            # Insert group headers in reverse order so indexing doesn't shift
            for (pos, cat_value_info) in zip(reversed(insert_positions), reversed(group_types)):
                _, cat_value = cat_value_info
                ws.insert_rows(pos, 1)
                header_cell = ws.cell(row=pos, column=1)
                header_cell.value = f"{cat_value}"
                header_cell.font = category_font
                header_cell.fill = fill
                header_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)

def process_file(file_path, output_directory, selected_brands):
    """
    Process a single CSV file, filtering to only the selected brands.
    Returns (unavailable_data, processed_file_base_name).
    """
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None, None

    existing_cols = [c for c in INPUT_COLUMNS if c in df.columns]
    if not existing_cols:
        print(f"[WARN] {file_path} is missing required columns. Skipped.")
        return None, None

    df = df[existing_cols]

    # Filter out 'promo' or 'sample'
    if 'Product' in df.columns:
        df = df[~df['Product'].str.contains(r'(?i)\bsample\b|\bpromo\b', na=False)]

    if 'Available' not in df.columns:
        print(f"[WARN] 'Available' not found in {file_path}, skipping.")
        return None, None

    unavailable_data = df[df['Available'] <= 2]
    available_data = df[df['Available'] > 2]

    # If we only want certain brands:
    if 'Brand' in available_data.columns and selected_brands:
        available_data = available_data[available_data['Brand'].isin(selected_brands)]

    # Extract strain and product details
    if 'Product' in available_data.columns:
        available_data['Strain_Type'] = available_data['Product'].apply(extract_strain_type)
        available_data[['Product_Weight','Product_SubType']] = available_data['Product'].apply(
            lambda x: pd.Series(extract_product_details(x))
        )
        # Filter out empty / numeric-only products
        available_data = available_data[~available_data['Product'].apply(is_empty_or_numbers)]
    else:
        available_data['Strain_Type'] = ""
        available_data['Product_Weight'] = ""
        available_data['Product_SubType'] = ""

    # Sort final data
    sort_cols = []
    if 'Category' in available_data.columns:
        sort_cols.append('Category')
    sort_cols.append('Strain_Type')
    sort_cols.append('Product_Weight')
    sort_cols.append('Product_SubType')
    if 'Product' in available_data.columns:
        sort_cols.append('Product')

    available_data.sort_values(by=sort_cols, inplace=True, na_position='last')

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    parts = base_name.split('_')
    store_name = parts[-1] if len(parts) > 1 else "Unknown"

    # Create subfolder for this CSV
    sub_out = os.path.join(output_directory, base_name)
    safe_makedirs(sub_out)

    today_str = datetime.datetime.now().strftime("%m-%d-%Y")

    if 'Brand' in available_data.columns:
        # Group by brand
        if available_data.empty:
            # If all data was filtered out
            out_xlsx = os.path.join(sub_out, f"{store_name}_{base_name}_{today_str}.xlsx")
            with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                available_data.to_excel(writer, index=False, sheet_name="Available")
                if not unavailable_data.empty:
                    unavailable_data.to_excel(writer, index=False, sheet_name="Unavailable")
            format_excel_file(out_xlsx)
            print(f"[INFO] Created {out_xlsx} (no brand data after filtering).")
        else:
            for brand_name, brand_data in available_data.groupby('Brand'):
                out_xlsx = os.path.join(sub_out, f"{store_name}_{brand_name}_{today_str}.xlsx")
                with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                    brand_data.to_excel(writer, index=False, sheet_name="Available")
                    if not unavailable_data.empty and 'Brand' in unavailable_data.columns:
                        brand_unavail = unavailable_data[unavailable_data['Brand'] == brand_name]
                        if not brand_unavail.empty:
                            brand_unavail.to_excel(writer, index=False, sheet_name="Unavailable")

                format_excel_file(out_xlsx)
                print(f"[INFO] Created {out_xlsx}")
    else:
        # No Brand column
        out_xlsx = os.path.join(sub_out, f"{store_name}_{base_name}_{today_str}.xlsx")
        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            available_data.to_excel(writer, index=False, sheet_name="Available")
            if not unavailable_data.empty:
                unavailable_data.to_excel(writer, index=False, sheet_name="Unavailable")
        format_excel_file(out_xlsx)
        print(f"[INFO] Created {out_xlsx}")

    return unavailable_data, base_name

def organize_by_brand(output_directory):
    """
    Moves XLSX files into subfolders named after the brand if their 
    filename is "<Store>_<Brand>_<MM-DD-YYYY>.xlsx".
    """
    pattern = re.compile(r"^(.*?)_(.*?)_(\d{2}-\d{2}-\d{4})\.xlsx$")

    for root, dirs, files in os.walk(output_directory):
        for f in files:
            if f.lower().endswith(".xlsx"):
                match = pattern.match(f)
                if match:
                    _, brand_name, _ = match.groups()
                    if os.path.basename(root) == brand_name:
                        continue
                    brand_folder = os.path.join(output_directory, brand_name)
                    safe_makedirs(brand_folder)

                    old_path = os.path.join(root, f)
                    new_path = os.path.join(brand_folder, f)
                    print(f"Moving {old_path} → {new_path}")
                    shutil.move(old_path, new_path)

def process_files(input_directory, output_directory, selected_brands):
    """
    Iterate all CSV files in `input_directory`, process them (filter by `selected_brands`),
    place XLSXs into `output_directory`. Then re-organize by brand subfolders.
    Returns a list of all final XLSX file paths.
    """
    safe_makedirs(output_directory)

    # Process each CSV
    for fn in os.listdir(input_directory):
        if fn.lower().endswith(".csv"):
            csv_path = os.path.join(input_directory, fn)
            try:
                process_file(csv_path, output_directory, selected_brands)
            except Exception as e:
                print(f"[ERROR] While processing {fn}: {e}")

    # Re-organize by brand
    organize_by_brand(output_directory)

    # Collect final XLSX
    final_files = []
    for root, dirs, files in os.walk(output_directory):
        for f in files:
            if f.lower().endswith(".xlsx"):
                final_files.append(os.path.join(root, f))

    return final_files


# ------------------------------------------------------------------------------
# --------------------------------- MAIN ---------------------------------------
# ------------------------------------------------------------------------------

def main():
    # 1) Clear out old CSVs from the "files" directory
    if os.path.exists(INPUT_DIRECTORY):
        for filename in os.listdir(INPUT_DIRECTORY):
            file_path = os.path.join(INPUT_DIRECTORY, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    print(f"[INFO] Deleted old CSV: {file_path}")
            except Exception as e:
                print(f"[ERROR] Could not delete {file_path}: {e}")

    # 2) Determine today's day name
    today_name = datetime.datetime.now().strftime("%A")  # e.g. "Monday", "Tuesday"

    # 3) Load brand_config.json
    if not os.path.exists(BRAND_CONFIG_JSON):
        print(f"[ERROR] Cannot find {BRAND_CONFIG_JSON}. Exiting.")
        sys.exit(1)

    with open(BRAND_CONFIG_JSON, "r", encoding="utf-8") as f:
        config = json.load(f)

    # read top-level test_mode, test_email
    test_mode = config.get("test_mode", True)
    test_email = config.get("test_email", "anthony@barbaro.tech")

    # read brand array
    brand_cfgs = config.get("brands", [])
    if not brand_cfgs:
        print("[INFO] No brand definitions found in brand_config.json -> 'brands' array.")
        return

    # ---------------------------------------------------------------
    # 4) Build a dictionary of brand synonyms -> (folder_name, emails)
    #    Also build brand_to_emails keyed by folder_name for emailing
    # ---------------------------------------------------------------
    synonym_to_folder = {}
    brand_to_emails = {}   # key = folder_name, value = final_emails

    for item in brand_cfgs:
        # brand_synonyms is a list of exact brand names from the CSV 'Brand' column
        synonyms = item.get("brand_synonyms", [])
        if isinstance(synonyms, str):
            synonyms = [synonyms]

        # fallback to old "brand" field if brand_synonyms is empty
        if not synonyms and "brand" in item:
            brand_str = item["brand"]
            synonyms = [b.strip() for b in brand_str.split('/')]

        folder_name = item.get("folder_name")
        if not folder_name:
            # if user didn't provide folder_name, fallback to first synonym
            folder_name = synonyms[0] if synonyms else "Unknown"

        real_emails = item.get("emails", [])
        days = item.get("days", [])
        location_str = item.get("location", "")  # optional reference

        # skip if not scheduled today
        if today_name not in days:
            continue

        # if test_mode => override emails
        final_emails = [test_email] if test_mode else real_emails

        # For each synonym brand name, map to (folder_name, final_emails)
        for syn in synonyms:
            synonym_to_folder[syn] = (folder_name, final_emails)

        # We'll store folder_name -> final_emails in brand_to_emails
        brand_to_emails[folder_name] = final_emails

    # If no folder_name is active, exit
    if not brand_to_emails:
        print(f"[INFO] No brands scheduled for {today_name}.")
        return

    # active_brands is the set of all "folder_name" keys from brand_to_emails
    active_brands = set(brand_to_emails.keys())

    print(f"[INFO] Today is {today_name}, active brand folders: {active_brands}")
    if test_mode:
        print(f"[INFO] TEST MODE ON => all emails go to {test_email}")

    # 5) Optionally call getCatalog.py
    try:
        print("[INFO] Running getCatalog.py to fetch latest CSV files ...")
        subprocess.check_call(["python", "getCatalog.py", INPUT_DIRECTORY])
        print("[INFO] CSV fetch complete.")
    except FileNotFoundError:
        print("[WARN] getCatalog.py not found, skipping CSV fetch step.")
    except subprocess.CalledProcessError as e:
        print(f"[ERROR] getCatalog.py failed: {e}")

    # ----------------------------------------------------------------
    # 6) synonyms_for_today => process CSV
    #    We pass ALL synonyms from synonym_to_folder
    # ----------------------------------------------------------------
    synonyms_for_today = list(synonym_to_folder.keys())
    safe_makedirs(LOCAL_REPORTS_FOLDER)
    generated_files = process_files(INPUT_DIRECTORY, LOCAL_REPORTS_FOLDER, synonyms_for_today)

    if not generated_files:
        print("[INFO] No XLSX files were generated. Possibly no data matched.")
        return

    # 7) Upload to Google Drive
    drive_service = drive_authenticate()
    parent_folder_id = find_or_create_folder(drive_service, DRIVE_PARENT_FOLDER_NAME, parent_id=None)
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    date_folder_id = find_or_create_folder(drive_service, date_str, parent_id=parent_folder_id)

    # For each folder_name in active_brands, create on Drive
    brand_folder_links = {}
    for folder_name in active_brands:
        brand_folder_id = find_or_create_folder(drive_service, folder_name, parent_id=date_folder_id)
        link = f"https://drive.google.com/drive/folders/{brand_folder_id}"
        brand_folder_links[folder_name] = link

    # Now, parse brand from each generated XLSX => find folder_name => upload
    brand_pattern = re.compile(r"^(.*?)_(.*?)_(\d{2}-\d{2}-\d{4})\.xlsx$", re.IGNORECASE)

    import time

    for file_path in generated_files:
        filename = os.path.basename(file_path)
        m = brand_pattern.match(filename)
        if not m:
            print(f"[WARN] Cannot parse brand from {filename}, skipping.")
            continue

        store_part, brand_syn, _ = m.groups()

        if brand_syn not in synonym_to_folder:
            print(f"[WARN] brand '{brand_syn}' not recognized. Skipping.")
            continue

        folder_name, _ = synonym_to_folder[brand_syn]
        if folder_name not in active_brands:
            continue

        # ✅ Reuse folder ID from earlier lookup
        if folder_name in brand_folder_links:
            brand_folder_id = brand_folder_links[folder_name].split("/")[-1]
        else:
            print(f"[ERROR] Missing folder ID for {folder_name}, skipping upload.")
            continue

        try:
            upload_file_to_drive(drive_service, file_path, brand_folder_id)
            print(f"[UPLOAD ✅] {filename} uploaded to {folder_name}")
            time.sleep(0.2)  # Google API throttle protection
        except Exception as e:
            print(f"[ERROR] Failed to upload {filename} → {folder_name}: {e}")

    # 8) Email out the folder link
    # Group by unique sets of emails
    email_groups = {}
    for folder_name, email_list in brand_to_emails.items():
        email_key = frozenset(email_list)
        if email_key not in email_groups:
            email_groups[email_key] = []
        email_groups[email_key].append(folder_name)

    for email_key, folder_list in email_groups.items():
        brand_lines = []
        for f_name in folder_list:
            link = brand_folder_links.get(f_name)
            if link:
                brand_lines.append(f"<h3>Folder: {f_name}</h3>")
                brand_lines.append(f"<p>Link: <a href='{link}'>{link}</a></p>")
            else:
                brand_lines.append(f"<p>No link found for {f_name}</p>")

        brand_html = "\n".join(brand_lines)
        subject = f"Brand Inventory Reports for {today_name} – {', '.join(folder_list)}"
        html_body = f"""
        <html>
        <body>
          <p>Hello,</p>
          <p>Below are your brand inventory reports for <strong>{today_name}</strong>.</p>
          {brand_html}
          <p>All files in that Drive folder are viewable by anyone with the link.</p>
          <p>Regards,<br>Buzz Cannabis</p>
        </body>
        </html>
        """

        recipients = list(email_key)
        print(f"[INFO] Sending Gmail API email to {recipients} for folders {folder_list} ...")
        send_email_with_gmail_html(subject, html_body, recipients)

    print("[INFO] All done!")

    # 9) Clean up
    if os.path.exists(LOCAL_REPORTS_FOLDER):
        try:
            shutil.rmtree(LOCAL_REPORTS_FOLDER)
            print(f"[INFO] Deleted temporary folder: {LOCAL_REPORTS_FOLDER}")
        except Exception as e:
            print(f"[ERROR] Could not delete {LOCAL_REPORTS_FOLDER}: {e}")



# ------------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[FATAL] Unhandled exception in BrandINVEmailer.py:")
        traceback.print_exc()
        sys.exit(1)
