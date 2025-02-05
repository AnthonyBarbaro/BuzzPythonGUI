import os
import re
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import traceback
from datetime import datetime

CONFIG_FILE = "config.txt"

INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand']
store_abbr_map = {
    "Buzz Cannabis - Mission Valley": "MV",
    "Buzz Cannabis-La Mesa": "LM",
    "Buzz Cannabis - SORRENTO VALLEY": "SV"
}
def ensure_dir_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

def extract_strain_type(product_name: str):
    if not isinstance(product_name, str):
        return ""
    name = " " + product_name.upper() + " "
    if re.search(r'\bS\b', name):
        return 'S'
    if re.search(r'\bH\b', name):
        return 'H'
    if re.search(r'\bI\b', name):
        return 'I'
    return ""

def extract_product_details(product_name: str):
    if not isinstance(product_name, str):
        return "", ""
    name_upper = product_name.upper()

    weight_match = re.search(r'(\d+(\.\d+)?)G', name_upper)
    weight = weight_match.group(0) if weight_match else ""
    sub_type = ""
    if " HH " in f" {name_upper} ":
        sub_type = "HH"
    elif " IN " in f" {name_upper} ":
        sub_type = "IN"

    return weight, sub_type

def is_empty_or_numbers(val):
    if not isinstance(val, str):
        return True
    val_str = val.strip()
    return val_str == "" or val_str.isdigit()

def format_excel_file(filename: str):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        available_col = None
        category_col = None
        strain_col = None
        product_col = None
        brand_col = None
        for i, cell in enumerate(ws[1], start=1):
            val = (cell.value or '').lower()
            if val == 'category':
                category_col = i
            elif val == 'available':
                available_col = i
            elif val == 'product':
                product_col = i
            elif val == 'brand':
                brand_col = i
            elif val == 'strain_type':
                strain_col = i

        if available_col:
            col_letter = get_column_letter(available_col)
            if ws.column_dimensions[col_letter].width < 20:
                ws.column_dimensions[col_letter].width = 20

        if category_col:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            current_type = None
            insert_positions = []
            for idx, row in enumerate(rows, start=2):
                pt = row[category_col - 1]
                if pt != current_type:
                    if current_type is not None:
                        insert_positions.append(idx)
                    current_type = pt
            if rows:
                insert_positions.insert(0, 2)

            category_font = Font(bold=True, size=14)
            fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            current_type = None
            group_types = []
            row_counter = 2
            for row_data in rows:
                pt = row_data[category_col - 1]
                if pt != current_type:
                    group_types.append((row_counter, pt))
                    current_type = pt
                row_counter += 1

            for (pos, pt_info) in zip(reversed(insert_positions), reversed(group_types)):
                pt_row, pt_value = pt_info
                ws.insert_rows(pos, 1)
                header_cell = ws.cell(row=pos, column=1)
                header_cell.value = f"Category: {pt_value}"
                header_cell.font = category_font
                header_cell.fill = fill
                max_col = ws.max_column
                ws.merge_cells(start_row=pos, start_column=1, end_row=pos, end_column=max_col)
                header_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)

def process_file(file_path, output_directory, selected_brands):
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None, None

    existing_cols = [c for c in INPUT_COLUMNS if c in df.columns]
    if not existing_cols:
        print(f"No required columns found in {file_path}. Skipping.")
        return None, None

    df = df[existing_cols]

    if 'Available' not in df.columns:
        print(f"'Available' column not found in {file_path}. Skipping.")
        return None, None

    unavailable_data = df[df['Available'] == 0]
    available_data = df[df['Available'] != 0]

    if 'Brand' in available_data.columns and selected_brands:
        available_data = available_data[available_data['Brand'].isin(selected_brands)]

    if 'Product' in available_data.columns:
        available_data['Strain_Type'] = available_data['Product'].apply(extract_strain_type)
        available_data[['Product_Weight','Product_SubType']] = available_data['Product'].apply(
            lambda x: pd.Series(extract_product_details(x))
        )
    else:
        available_data['Strain_Type'] = ""
        available_data['Product_Weight'] = ""
        available_data['Product_SubType'] = ""

    if 'Product' in available_data.columns:
        available_data = available_data[~available_data['Product'].apply(is_empty_or_numbers)]

    sort_cols = []
    if 'Category' in available_data.columns:
        sort_cols.append('Category')
    sort_cols.append('Strain_Type')
    sort_cols.append('Product_Weight')
    sort_cols.append('Product_SubType')
    if 'Product' in available_data.columns:
        sort_cols.append('Product')

    available_data.sort_values(by=sort_cols, inplace=True, na_position='last')
# Extract store name from file name (assumes last part before ".csv" is the store name)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    parts = base_name.split('_')
    store_name = parts[-1] if len(parts) > 1 else "Unknown"

    # Add Store Name Column
    available_data.insert(0, 'Store', store_name)
    unavailable_data.insert(0, 'Store', store_name)

    file_subdir = os.path.join(output_directory, base_name)
    ensure_dir_exists(file_subdir)
    today_str = datetime.now().strftime("%m-%d-%Y")

    brand_exists = 'Brand' in available_data.columns
    if brand_exists:
        if available_data.empty:
            output_filename = os.path.join(file_subdir, f"{store_name}_{base_name}_{today_str}.xlsx")
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                available_data.to_excel(writer, index=False, sheet_name='Available')
                if not unavailable_data.empty:
                    unavailable_data.to_excel(writer, index=False, sheet_name='Unavailable')
            format_excel_file(output_filename)
            print(f"Created {output_filename} (no brand data after filtering)")
        else:
            for brand, brand_data in available_data.groupby('Brand'):
                output_filename = os.path.join(file_subdir, f"{store_name}_{brand}_{today_str}.xlsx")
                with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                    brand_data.to_excel(writer, index=False, sheet_name='Available')
                    if not unavailable_data.empty:
                        if 'Brand' in unavailable_data.columns:
                            brand_unavail = unavailable_data[unavailable_data['Brand'] == brand]
                        else:
                            brand_unavail = pd.DataFrame()
                        if not brand_unavail.empty:
                            brand_unavail.to_excel(writer, index=False, sheet_name='Unavailable')
                format_excel_file(output_filename)
                print(f"Created {output_filename}")
    else:
        output_filename = os.path.join(file_subdir, f"{store_name}_{base_name}_{today_str}.xlsx")
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            available_data.to_excel(writer, index=False, sheet_name='Available')
            if not unavailable_data.empty:
                unavailable_data.to_excel(writer, index=False, sheet_name='Unavailable')
        format_excel_file(output_filename)
        print(f"Created {output_filename}")

    return unavailable_data, os.path.basename(file_path)

def process_files(input_directory, output_directory, selected_brands):
    ensure_dir_exists(output_directory)

    summary_file = os.path.join(output_directory, 'done.csv')
    unavailable_file = os.path.join(output_directory, 'unavailable.csv')

    if os.path.exists(summary_file):
        summary_df = pd.read_csv(summary_file)
    else:
        summary_df = pd.DataFrame(columns=['File', 'Status'])

    if os.path.exists(unavailable_file):
        unavailable_df = pd.read_csv(unavailable_file)
    else:
        unavailable_df = pd.DataFrame()

    for filename in os.listdir(input_directory):
        if filename.endswith('.csv'):
            file_path = os.path.join(input_directory, filename)
            try:
                unavail_data, processed_file = process_file(file_path, output_directory, selected_brands)
                if processed_file is not None:
                    if unavail_data is not None and not unavail_data.empty:
                        if 'Source File' not in unavail_data.columns:
                            unavail_data['Source File'] = processed_file
                        unavailable_df = pd.concat([unavailable_df, unavail_data], ignore_index=True)
                    summary_df = pd.concat([summary_df, pd.DataFrame({
                        'File': [processed_file],
                        'Status': ["Processed successfully"]
                    })], ignore_index=True)
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                summary_df = pd.concat([summary_df, pd.DataFrame({
                    'File': [filename],
                    'Status': [f"Error: {str(e)}"]
                })], ignore_index=True)

    summary_df.to_csv(summary_file, index=False)
    print(f"Summary results saved to {summary_file}")

    unavailable_df.to_csv(unavailable_file, index=False)
    print(f"Unavailable products saved to {unavailable_file}")

    if os.path.exists(summary_file):
        os.remove(summary_file)
        print("Deleted done.csv")
    if os.path.exists(unavailable_file):
        os.remove(unavailable_file)
        print("Deleted unavailable.csv")

def get_all_brands(input_directory):
    brands = set()
    brand_found = False
    for filename in os.listdir(input_directory):
        if filename.endswith('.csv'):
            file_path = os.path.join(input_directory, filename)
            try:
                df = pd.read_csv(file_path)
                if 'Brand' in df.columns:
                    brand_found = True
                    new_brands = df['Brand'].dropna().unique().tolist()
                    brands.update(new_brands)
            except:
                pass
    
    if not brand_found:
        return []
    return sorted(list(brands))

def save_config(input_dir, output_dir):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        f.write(f"{input_dir}\n")
        f.write(f"{output_dir}\n")

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            lines = f.read().strip().split('\n')
            if len(lines) >= 2:
                input_dir = lines[0].strip()
                output_dir = lines[1].strip()
                if os.path.isdir(input_dir) and os.path.isdir(output_dir):
                    return input_dir, output_dir
    return None, None

def auto_detect_dirs():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    files_dir = os.path.join(script_dir, 'files')
    done_dir = os.path.join(script_dir, 'done')
    if os.path.isdir(files_dir) and os.path.isdir(done_dir):
        return files_dir, done_dir
    return None, None

class BrandInventoryApp:
    """Brand Inventory Report application screen."""
    def __init__(self, master, return_to_main):
        self.master = master
        self.return_to_main = return_to_main

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()

        i_dir, o_dir = load_config()

        if not i_dir or not o_dir:
            i_dir, o_dir = auto_detect_dirs()

        if i_dir and o_dir:
            self.input_dir.set(i_dir)
            self.output_dir.set(o_dir)

        frame = tk.Frame(master)
        self.frame = frame
        frame.pack(pady=10)

        # Directory selection
        dir_frame = tk.Frame(frame)
        dir_frame.pack(pady=10)
        
        tk.Label(dir_frame, text="Input Directory:").grid(row=0, column=0, sticky='e')
        tk.Entry(dir_frame, textvariable=self.input_dir, width=50).grid(row=0, column=1)
        tk.Button(dir_frame, text="Browse", command=self.browse_input).grid(row=0, column=2, padx=5)

        tk.Label(dir_frame, text="Output Directory:").grid(row=1, column=0, sticky='e')
        tk.Entry(dir_frame, textvariable=self.output_dir, width=50).grid(row=1, column=1)
        tk.Button(dir_frame, text="Browse", command=self.browse_output).grid(row=1, column=2, padx=5)

        # Brand selection
        brand_frame = tk.Frame(frame)
        brand_frame.pack(pady=10)

        tk.Label(brand_frame, text="Select Brands (Ctrl+Click to select multiple):").pack(anchor='w')

        self.brand_listbox = tk.Listbox(brand_frame, selectmode=tk.MULTIPLE, height=10, width=50)
        self.brand_listbox.pack(pady=5)

        # Buttons
        btn_frame = tk.Frame(frame)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Get Files", command=self.get_files).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Clear Files", command=self.clear_files).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Load Brands", command=self.load_brands).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Select All", command=self.select_all_brands).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Generate Reports", command=self.run_process).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Return to Main Hub", command=self.return_to_main_hub).pack(side='left', padx=5)

    def return_to_main_hub(self):
        self.frame.destroy()
        self.return_to_main()

    def get_files(self):
        input_dir = self.input_dir.get()
        if not input_dir:
            messagebox.showerror("Error", "Please select an input directory first.")
            return
        try:
            subprocess.check_call(["python", "getCatalog.py", input_dir])
            messagebox.showinfo("Success", "Files successfully fetched.")
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error", f"Failed to get files: {e}")

    def clear_files(self):
        input_dir = self.input_dir.get()
        if not input_dir:
            messagebox.showerror("Error", "Please select an input directory first.")
            return

        # Confirm the deletion
        answer = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete all CSV files in the input directory?")
        if not answer:
            return

        count = 0
        for filename in os.listdir(input_dir):
            if filename.endswith('.csv'):
                file_path = os.path.join(input_dir, filename)
                try:
                    os.remove(file_path)
                    count += 1
                except Exception as e:
                    print(f"Error deleting {filename}: {e}")
        messagebox.showinfo("Files Deleted", f"Deleted {count} CSV files from {input_dir}.")

    def browse_input(self):
        directory = filedialog.askdirectory()
        if directory:
            self.input_dir.set(directory)

    def browse_output(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)

    def load_brands(self):
        if not self.input_dir.get():
            messagebox.showerror("Error", "Please select an input directory first.")
            return
        try:
            brands = get_all_brands(self.input_dir.get())
            self.brand_listbox.delete(0, tk.END)
            self.brand_listbox.configure(state='normal')
            if brands is None or len(brands) == 0:
                self.brand_listbox.insert(tk.END, "No brands found. You can still run the report.")
            else:
                for b in brands:
                    self.brand_listbox.insert(tk.END, b)
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load brands: {e}")

    def select_all_brands(self):
        if self.brand_listbox['state'] == 'normal':
            if self.brand_listbox.size() > 0:
                first_item = self.brand_listbox.get(0)
                if "No brands found" in first_item:
                    messagebox.showinfo("Info", "No actual brands available to select.")
                else:
                    self.brand_listbox.select_set(0, tk.END)
            else:
                messagebox.showinfo("Info", "No brands available to select.")
        else:
            messagebox.showinfo("Info", "No brands to select.")

    def run_process(self):
        input_dir = self.input_dir.get()
        output_dir = self.output_dir.get()
        if not input_dir or not output_dir:
            messagebox.showerror("Error", "Please select both input and output directories.")
            return

        selected_brands = []
        if self.brand_listbox['state'] == 'normal':
            selected_indices = self.brand_listbox.curselection()
            selected_brands = [self.brand_listbox.get(i) for i in selected_indices]

        try:
            process_files(
                input_directory=input_dir,
                output_directory=output_dir,
                selected_brands=selected_brands
            )
            save_config(input_dir, output_dir)
            messagebox.showinfo("Success", "Reports generated successfully.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Error generating reports:\n{e}")

class MainHub:
    """Main hub with multiple app choices."""
    def __init__(self, master):
        self.master = master
        frame = tk.Frame(master)
        self.frame = frame
        frame.pack(pady=20)

        tk.Label(frame, text="Main Hub", font=("Arial", 16, "bold")).pack(pady=10)

        tk.Button(frame, text="Brand Inventory Report", command=self.open_brand_inventory).pack(pady=5)
        tk.Button(frame, text="Sales Area (Placeholder)", command=self.sales_area).pack(pady=5)
        tk.Button(frame, text="Another Feature (Placeholder)", command=self.another_feature).pack(pady=5)
        tk.Button(frame, text="Exit", command=self.exit_app).pack(pady=5)

    def open_brand_inventory(self):
        self.frame.destroy()
        BrandInventoryApp(self.master, self.return_to_main)

    def sales_area(self):
        messagebox.showinfo("Info", "Sales Area is not implemented yet.")

    def another_feature(self):
        messagebox.showinfo("Info", "Another feature is not implemented yet.")

    def exit_app(self):
        self.master.quit()

    def return_to_main(self):
        MainHub(self.master)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Main Hub - Multiple Apps")
    root.attributes('-topmost', True)
    MainHub(root)
    root.mainloop()