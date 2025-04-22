#!/usr/bin/env python3

import os
import re
import openpyxl

# ---------------------------------------------
# 1) GMAIL API SEND LOGIC 
# ---------------------------------------------
def send_email_with_gmail_html(subject, html_body, recipients, attachments=None):
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    gmail_token = "token_gmail.json"

    creds = None
    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token, "w") as f:
            f.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    msg = MIMEMultipart('alternative')
    msg['From'] = "me"
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    part_html = MIMEText(html_body, 'html')
    msg.attach(part_html)

    if attachments:
        for file_path in attachments:
            if not os.path.isfile(file_path):
                continue
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as fp:
                file_data = fp.read()
            part = MIMEBase("application", "octet-stream")
            part.set_payload(file_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_req = {'raw': raw_message}

    try:
        sent = service.users().messages().send(userId='me', body=send_req).execute()
        print(f"[INFO] HTML Email sent! ID: {sent['id']} | Subject: {subject}")
    except Exception as e:
        print(f"[ERROR] Could not send HTML email: {e}")


# ---------------------------------------------
# 2) BRAND -> EMAIL MAP (adjust for real usage)
# ---------------------------------------------
BRAND_EMAILS = {
   # e.g. "Heavy Hitters": "heavyhittersteam@example.com",
   # "WYLD GoodTide": "wyldteam@example.com",
}
DEFAULT_EMAIL = "anthony@barbaro.tech"


# ---------------------------------------------
# 3) PARSE "SUMMARY" SHEET FOR (STORE, KICKBACK OWED)
# ---------------------------------------------
def parse_kickback_summary(brand_report_path):
    """Returns list[(store, owed)], ignoring 'Store'/'Kickback Owed' placeholders."""
    results = []
    if not os.path.isfile(brand_report_path):
        return results

    wb = openpyxl.load_workbook(brand_report_path, data_only=True)
    if "Summary" not in wb.sheetnames:
        wb.close()
        return results

    sh = wb["Summary"]
    for row_idx in range(2, sh.max_row + 1):
        store_val = sh.cell(row=row_idx, column=1).value
        owed_val  = sh.cell(row=row_idx, column=2).value
        if store_val is not None and owed_val is not None:
            s_str = str(store_val).strip().lower()
            o_str = str(owed_val).strip().lower()
            if s_str not in ["store", ""] and o_str not in ["kickback owed", ""]:
                results.append((store_val, owed_val))
    wb.close()
    return results


def build_kickback_table(rows):
    if not rows:
        return "<p>(No data found in summary.)</p>"

    html = """
    <table border='1' cellpadding='5' cellspacing='0'>
      <thead>
        <tr><th>Store</th><th>Kickback Owed</th></tr>
      </thead>
      <tbody>
    """
    for (store, owed) in rows:
        try:
            owed_float = float(owed)
            owed_str = f"${owed_float:,.2f}"
        except:
            owed_str = str(owed)
        html += f"<tr><td>{store}</td><td>{owed_str}</td></tr>"
    html += "</tbody></table>"
    return html


# ---------------------------------------------
# 4) PARSE links.txt FOR BRAND LINKS
# ---------------------------------------------
def parse_brand_and_link(line):
    """
    Attempt to parse brand name from filename forms like:
      - "Heavy Hitters_report_2025-02-21_to_2025-02-22.xlsx: <URL>"
      - "WYLD GoodTide_report_2025-02-21_to_2025-02-22.xlsx: <URL>"
      - "Hashish_02-24-2025_SV.xlsx: <URL>"
    capturing everything (including spaces) up to `_report_` 
    or up to the date pattern.
    """
    if ':' not in line:
        return None, None
    left_part, url_part = line.split(':', 1)
    left_part = left_part.strip()
    url_part  = url_part.strip()
    if not url_part.startswith("http"):
        return None, None

    # Pattern A: brand up to '_report_'
    m = re.match(r'^(.+?)_report_', left_part, re.IGNORECASE)
    if m:
        brand = m.group(1).strip()
        return brand, line

    # Pattern B: brand up to '_' + date
    # e.g. "Hashish_02-24-2025_SV.xlsx"
    # We'll capture everything before the date:
    m2 = re.match(r'^(.+?)_[0-9]{2}-[0-9]{2}-[0-9]{4}', left_part, re.IGNORECASE)
    if m2:
        brand = m2.group(1).strip()
        return brand, line

    return None, None


def load_links_map(links_file="links.txt"):
    """
    Returns { normalized_brand: [full_line_1, full_line_2, ...], ... }
    """
    brand_links_map = {}
    if not os.path.isfile(links_file):
        print(f"[WARN] no links.txt found at {links_file}.")
        return brand_links_map

    with open(links_file, "r", encoding="utf-8") as f:
        lines = [ln.strip() for ln in f if ln.strip()]

    for line in lines:
        brand, full_line = parse_brand_and_link(line)
        if brand:
            # e.g. "Heavy Hitters" -> "heavyhitters"
            normalized = brand.lower().replace(" ", "")
            if normalized not in brand_links_map:
                brand_links_map[normalized] = []
            brand_links_map[normalized].append(line)
        else:
            # If we can't parse brand, we skip or store in a "misc" bucket
            pass

    return brand_links_map


def make_html_link_list(lines):
    """
    Convert lines like:
      "Heavy Hitters_report_2025-02-21_to_2025-02-22.xlsx: https://..."
    into an HTML <ul> with clickable links.
    """
    if not lines:
        return "<p>No links for this brand.</p>"
    html = "<ul>"
    for ln in lines:
        if ":" in ln:
            filename, link = ln.split(":", 1)
            filename = filename.strip()
            link = link.strip()
            html += f"<li><strong>{filename}</strong>: <a href='{link}'>{link}</a></li>"
        else:
            html += f"<li>{ln}</li>"
    html += "</ul>"
    return html


# ---------------------------------------------
# 5) MAIN "SEND BRAND EMAILS" LOGIC
# ---------------------------------------------
def send_brand_emails():
    reports_dir = "brand_reports"
    if not os.path.isdir(reports_dir):
        print(f"[ERROR] The folder '{reports_dir}' does not exist. No emails sent.")
        return

    # 1) Load brand->linklines from links.txt
    brand_links_map = load_links_map("links.txt")

    all_brands_info = []

    # 2) Iterate brand_reports, parse brand, build email
    for filename in os.listdir(reports_dir):
        if not filename.endswith(".xlsx"):
            continue

        # Pattern A: brand up to '_report_'
        match = re.match(r'^(.+?)_report_', filename, re.IGNORECASE)
        if match:
            brand_name = match.group(1).strip()
        else:
            # Pattern B: brand up to '_dd-mm-yyyy'
            match2 = re.match(r'^(.+?)_[0-9]{2}-[0-9]{2}-[0-9]{4}', filename, re.IGNORECASE)
            if match2:
                brand_name = match2.group(1).strip()
            else:
                brand_name = filename.rsplit('.', 1)[0]  # fallback entire base

        file_path = os.path.join(reports_dir, filename)

        # 3) Parse Kickback summary
        summary_rows = parse_kickback_summary(file_path)
        owed_table_html = build_kickback_table(summary_rows)

        # 4) Gather brand Drive links from brand_links_map
        normalized_brand = brand_name.lower().replace(" ", "")
        brand_link_lines = brand_links_map.get(normalized_brand, [])
        link_list_html = make_html_link_list(brand_link_lines)

        # 5) Determine recipient
        recipient = [BRAND_EMAILS.get(brand_name, DEFAULT_EMAIL)]
        #recipient = [BRAND_EMAILS.get(brand_name, DEFAULT_EMAIL)]


        # 6) Build HTML body
        html_body = f"""
        <html>
          <body>
            <p>Hello {brand_name},</p>
            <p>Please see below your Kickback Owed details and the Drive link(s):</p>
            <h3>Kickback Summary:</h3>
            {owed_table_html}
            <h3>Google Drive Link(s):</h3>
            {link_list_html}
            <p>Attached is your brand sales report as well, if needed.</p>
            <p>Regards,<br>Anthony</p>
            <p><strong>please include/contact anthony@buzzcannabis.com & donna@buzzcannabis.com in all emails regarding these credits. </strong></p>
          </body>
        </html>
        """

        subject = f"Weekly Kickback - {brand_name}"
        print(f"[INFO] Sending email for brand '{brand_name}' to '{recipient}' ...")
        send_email_with_gmail_html(
            subject=subject,
            html_body=html_body,
            recipients=recipient,
            attachments=[file_path]
        )

        # Collect info for optional consolidated
        total_owed = 0.0
        for (_, o) in summary_rows:
            try:
                total_owed += float(o)
            except:
                pass
        all_brands_info.append({
            "brand": brand_name,
            "rows": summary_rows,
            "links": brand_link_lines,
            "total_owed": total_owed
        })

    # 7) OPTIONAL: Send consolidated email
    if all_brands_info:
        grand_total = sum(bi["total_owed"] for bi in all_brands_info)
        brand_htmls = []
        for bi in all_brands_info:
            brand = bi["brand"]
            link_html = make_html_link_list(bi["links"])
            row_html = build_kickback_table(bi["rows"])
            brand_html = f"""
            <h3>{brand}</h3>
            <p><strong>Links:</strong></p>
            {link_html}
            <p><strong>Kickback Summary:</strong></p>
            {row_html}
            """
            brand_htmls.append(brand_html)

        consolidated_html = "<hr>".join(brand_htmls)
        consolidated_html += f"<p><strong>GRAND TOTAL OWED: ${grand_total:,.2f}</strong></p>"

        subject = "Consolidated Weekly Kickback (All Brands)"
        html_body = f"""
        <html>
        <body>
          <p>Hello Team,</p>
          <p>Below is a consolidated summary of all brands Kickback info:</p>
          {consolidated_html}
          <p>Regards,<br>Donna</p>
        <p><strong>please include/contact anthony@buzzcannabis.com & donna@buzzcannabis.com in all emails regarding these credits. </strong></p>
        </body>
        </html>
        """

        send_email_with_gmail_html(
            subject=subject,
            html_body=html_body,
            recipients=["anthony@barbaro.tech","donna@buzzcannabis.com"],
            #recipients=["anthony@barbaro.tech"],
            attachments=None
        )


# ---------------------------------------------
# 6) MAIN
# ---------------------------------------------
if __name__ == "__main__":
    send_brand_emails()
