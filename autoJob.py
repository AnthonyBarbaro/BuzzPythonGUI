#!/usr/bin/env python3

import os
import re
import subprocess
import time
import traceback
import datetime
import calendar
from datetime import date, timedelta, datetime as dt
from pathlib import Path

##############################################################################
# 1) LOGIC FOR LAST MONDAY TO SUNDAY
##############################################################################
def get_last_monday_sunday():
    """
    Returns (start_date, end_date) as Python date objects, representing
    last Monday through Sunday.

    Example: if today is Monday 2025-01-20, 
    this returns Monday 2025-01-13 and Sunday 2025-01-19.
    """
    today = date.today()
    # Monday of THIS current week:
    monday_this_week = today - timedelta(days=today.weekday())
    # Last Monday is 7 days before the Monday of this week
    last_monday = monday_this_week - timedelta(days=7)
    # Last Sunday is last_monday + 6 days
    last_sunday = last_monday + timedelta(days=6)
    return last_monday, last_sunday


##############################################################################
# 2) GETCATALOG LOGIC
##############################################################################
def run_get_catalog():
    """
    Calls getCatalog.py with a simple subprocess. 
    Make sure getCatalog.py is in the same directory or specify the full path.
    """
    print("\n===== Running getCatalog.py to download Catalog files... =====\n")
    try:
        subprocess.check_call(["python", "getCatalog.py"])
    except subprocess.CalledProcessError as e:
        print(f"[ERROR] getCatalog.py failed: {e}")
    except FileNotFoundError:
        print("[ERROR] getCatalog.py not found. Please check the script name/path.")


##############################################################################
# 3) GETSALESREPORT LOGIC (HEADLESS, NO GUI)
#    We replicate the essential parts of your getSalesReport.py but skip GUI.
##############################################################################

#in salesReport.py


##############################################################################
# 4) RUN deals.py (Brand-Level Deals Report)
##############################################################################

#in deals.py

##############################################################################
# 5) RUN BRAND_INVENTORY.PY FOR 'Hashish' ONLY
##############################################################################
def run_brand_inventory_hashish():
    """
    We'll replicate a minimal version of brand_inventory.py logic,
    forcing brand='Hashish' only.
    We'll assume we want to parse the 'files' directory for new CSVs,
    output to 'done', and only keep lines for brand 'Hashish'.
    Then we apply openpyxl formatting (freeze panes, column widths, row height).
    """
    print("\n===== Running brand_inventory.py logic ONLY for brand='Hashish'... =====\n")

    import pandas as pd
    import re
    from datetime import datetime as dt
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    
    input_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "files")
    output_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "done")
    os.makedirs(output_directory, exist_ok=True)

    def is_empty_or_numbers(val):
        if not isinstance(val, str):
            return True
        val_str = val.strip()
        return val_str == "" or val_str.isdigit()

    def extract_strain_type(product_name: str):
        if not isinstance(product_name, str):
            return ""
        name = " " + product_name.upper() + " "
        if re.search(r'\bS\b', name):
            return 'S'
        if re.search(r'\bH\b', name):
            return 'H'
        if re.search(r'\bI\b', name):
            return 'I'
        return ""

    def extract_product_details(product_name: str):
        if not isinstance(product_name, str):
            return "", ""
        name_upper = product_name.upper()
        weight_match = re.search(r'(\d+(\.\d+)?)G', name_upper)
        weight = weight_match.group(0) if weight_match else ""
        sub_type = ""
        if " HH " in f" {name_upper} ":
            sub_type = "HH"
        elif " IN " in f" {name_upper} ":
            sub_type = "IN"
        return weight, sub_type

    INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand']

    for filename in os.listdir(input_directory):
        if filename.lower().endswith('.csv'):
            file_path = os.path.join(input_directory, filename)
            try:
                df = pd.read_csv(file_path)
            except Exception as e:
                print(f"[ERROR] reading CSV {filename}: {e}")
                continue

            # Filter to required columns
            use_cols = [c for c in INPUT_COLUMNS if c in df.columns]
            if not use_cols:
                continue
            df = df[use_cols]

            # Only brand=Hashish
            if 'Brand' in df.columns:
                df = df[df['Brand'] == 'Hashish']

            # Separate available vs. unavailable
            if 'Available' not in df.columns:
                continue
            unavailable_data = df[df['Available'] == 0]
            available_data   = df[df['Available'] != 0]

            # Parse product columns for the 'available' subset
            if not available_data.empty and 'Product' in available_data.columns:
                available_data['Strain_Type'] = available_data['Product'].apply(extract_strain_type)
                available_data[['Product_Weight','Product_SubType']] = available_data['Product'].apply(
                    lambda x: pd.Series(extract_product_details(x))
                )
                # Remove rows with empty or numeric product name
                available_data = available_data[~available_data['Product'].apply(is_empty_or_numbers)]

            # Sort by Category, Strain_Type, Product_Weight, Product_SubType, and Product
            sort_cols = []
            if 'Category' in available_data.columns:
                sort_cols.append('Category')
            sort_cols += ['Strain_Type','Product_Weight','Product_SubType']
            if 'Product' in available_data.columns:
                sort_cols.append('Product')
            available_data.sort_values(by=sort_cols, inplace=True, na_position='last')

            # Prepare output path
            base_name = os.path.splitext(filename)[0]  # e.g. "GreenHalo"
            today_str = dt.now().strftime("%m-%d-%Y")
            out_subdir = os.path.join(output_directory, base_name)
            os.makedirs(out_subdir, exist_ok=True)

            # --- Construct final Excel filename with "Hashish" + "_" + <inputfile base name> ---
            # e.g. "Hashish_GreenHalo.xlsx"
            out_file = os.path.join(out_subdir, f"Hashish_{base_name}.xlsx")

            # Write to Excel using pandas
            with pd.ExcelWriter(out_file) as writer:
                available_data.to_excel(writer, index=False, sheet_name='Available')
                if not unavailable_data.empty:
                    unavailable_data.to_excel(writer, index=False, sheet_name='Unavailable')

            # Apply formatting with openpyxl
            workbook = load_workbook(out_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Freeze the first row
                sheet.freeze_panes = "A2"

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 
                                     for cell in column)
                    sheet.column_dimensions[column[0].column_letter].width = max_length + 2

                # Set a default row height
                for row in sheet.iter_rows():
                    sheet.row_dimensions[row[0].row].height = 17

                # Make the first row bold & center-aligned
                for cell in sheet["1:1"]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            workbook.save(out_file)
            print(f"Hashish brand inventory saved & formatted -> {out_file}")


##############################################################################
# 6) GOOGLE DRIVE UPLOADER
##############################################################################
def run_drive_upload():
    """
    Upload brand_reports/*.xlsx + any done/**/*Hashish_*.xlsx to 
    Google Drive folder "2025_Kickback -> <week range>", 
    writing all links into links.txt
    """
    print("\n===== Running googleDriveUploader logic... =====\n")
    import google.auth
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    from google.oauth2.credentials import Credentials

    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    LINKS_FILE = "links.txt"
    PARENT_FOLDER_NAME = "2025_Kickback"
    REPORTS_FOLDER = "brand_reports"

    def authenticate_drive_api():
        creds = None
        token_file = "token.json"
        if os.path.exists(token_file):
            creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(google.auth.transport.requests.Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_file, "w") as token:
                token.write(creds.to_json())
        return build("drive","v3", credentials=creds)

    def get_week_range_str():
        lm, ls = get_last_monday_sunday()
        return f"{lm.strftime('%b %d')} to {ls.strftime('%b %d')}"

    def find_or_create_folder(service, folder_name, parent_id=None):
        query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}'"
        if parent_id:
            query += f" and '{parent_id}' in parents"
        resp = service.files().list(q=query, spaces='drive', fields='files(id,name)').execute()
        items = resp.get('files', [])
        if items:
            return items[0]['id']
        else:
            meta = {
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder"
            }
            if parent_id:
                meta["parents"] = [parent_id]
            f = service.files().create(body=meta, fields="id").execute()
            return f["id"]

    def upload_file(service, path, parent_id):
        fname = os.path.basename(path)
        body = {
            "name": fname,
            "parents": [parent_id]
        }
        media = MediaFileUpload(path, resumable=True)
        f = service.files().create(body=body, media_body=media, fields="id").execute()
        return f["id"]

    def make_public(service, file_id):
        try:
            perm = {"type":"anyone","role":"reader"}
            service.permissions().create(fileId=file_id, body=perm).execute()
            info = service.files().get(fileId=file_id, fields="webViewLink").execute()
            return info.get("webViewLink")
        except:
            return None

    service = authenticate_drive_api()
    parent_id = find_or_create_folder(service, PARENT_FOLDER_NAME, None)

    week_range = get_week_range_str()
    week_folder_id = find_or_create_folder(service, week_range, parent_id)

    with open(LINKS_FILE,"w") as lf:
        # 1) Upload brand_reports
        if os.path.isdir(REPORTS_FOLDER):
            for fname in os.listdir(REPORTS_FOLDER):
                if fname.endswith(".xlsx"):
                    full_path = os.path.join(REPORTS_FOLDER,fname)
                    # Upload to the *same* week folder (no sub-subfolders)
                    file_id = upload_file(service, full_path, week_folder_id)
                    link = make_public(service, file_id)
                    if link:
                        lf.write(f"{fname}: {link}\n")
                        print(f"Uploaded {fname} => {link}")

        # 2) Also upload done/**/*Hashish_*.xlsx
        done_dir = "done"
        if os.path.isdir(done_dir):
            for root, dirs, files in os.walk(done_dir):
                for f in files:
                    if f.endswith(".xlsx") and "Hashish_" in f:
                        full_path = os.path.join(root,f)
                        file_id = upload_file(service, full_path, week_folder_id)
                        link = make_public(service, file_id)
                        if link:
                            lf.write(f"{f}: {link}\n")
                            print(f"Uploaded {f} => {link}")

    print("All files uploaded. Links stored in links.txt.")


##############################################################################
# 7) EMAIL THE LINKS.TXT + HASHISH BRAND REPORT (Optional)
##############################################################################
def send_email_with_gmail(subject, body, recipients, attachments=None):
    """
    Sends an email (plain text) via Gmail API with optional attachments.
    """
    print("\n===== Sending Email via Gmail API... =====\n")
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

    creds = None
    gmail_token = "token_gmail.json"
    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token,"w") as t:
            t.write(creds.to_json())

    service = build('gmail','v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    msg = MIMEMultipart()
    msg['From'] = "me"
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    if attachments:
        for path in attachments:
            if not os.path.isfile(path):
                continue
            fn = os.path.basename(path)
            with open(path,"rb") as f:
                part = MIMEBase("application","octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{fn}"')
            msg.attach(part)

    raw_msg = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_msg = {'raw': raw_msg}
    try:
        sent = service.users().messages().send(userId='me', body=send_msg).execute()
        print(f"Email sent! ID: {sent['id']}")
    except Exception as e:
        print("[ERROR] Could not send Gmail:", e)
def send_email_with_gmail_html(subject, html_body, recipients, attachments=None):
    """
    Sends an HTML email via Gmail API with optional attachments.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    creds = None
    gmail_token = "token_gmail.json"

    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token, "w") as f:
            f.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    # Create a MIMEMultipart message for HTML
    msg = MIMEMultipart('alternative')
    msg['From'] = "me"
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    # Attach the HTML body
    part_html = MIMEText(html_body, 'html')
    msg.attach(part_html)

    # Optionally attach files
    if attachments:
        for file_path in attachments:
            if not os.path.isfile(file_path):
                continue
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as fp:
                file_data = fp.read()
            part = MIMEBase("application", "octet-stream")
            part.set_payload(file_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_req = {'raw': raw_message}

    try:
        sent = service.users().messages().send(userId='me', body=send_req).execute()
        print(f"HTML Email sent! ID: {sent['id']}")
    except Exception as e:
        print("[ERROR] Could not send HTML email:", e)

from getSalesReport import run_sales_report
##############################################################################
# MAIN: ORCHESTRATE ALL STEPS
##############################################################################
def main():
    print("===== Starting autoJob.py =====")

    last_monday, last_sunday = get_last_monday_sunday()
    date_range_str = f"{last_monday} to {last_sunday}"
    print(f"Processing for last week range: {date_range_str}")

    # 1) Catalog

    # 2) Sales
    run_sales_report(last_monday, last_sunday)

    # 3) Deals
    subprocess.run(["python", "deals.py"])
    time.sleep(2)

    # 5) Drive Upload (both brand_reports + done/Hashish)
    run_drive_upload()


    # 6a) Parse links.txt to separate "Hashish" lines from "non-Hashish" lines,
    #     and build HTML bullet lists for each group.
   

    links_file = "links.txt"
    hashish_links = []
    non_hashish_links = []

    if os.path.exists(links_file):
        with open(links_file, "r", encoding="utf-8") as lf:
            lines = lf.readlines()
        for line in lines:
            line = line.strip()
            # Typical format: "filename.xlsx: https://drive.google.com/..."
            if "Hashish_" in line:
                hashish_links.append(line)
            else:
                non_hashish_links.append(line)

    subprocess.run(["python", "brandDEALSEmailer.py"])
        # 7) Clean up files directory
    files_dir = Path("files")
    if files_dir.exists() and files_dir.is_dir():
        for file in files_dir.iterdir():
            try:
                if file.is_file():
                    file.unlink()
                    print(f"[CLEANUP] Deleted {file}")
            except Exception as e:
                print(f"[ERROR] Could not delete {file}: {e}")

    print("\n===== autoJob.py completed successfully. =====")


if __name__ == "__main__":
    main()