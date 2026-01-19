#!/usr/bin/env python3
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import locale
import shutil
import warnings
# Global dictionary to map real names -> pseudonyms
NAME_MAP = {}
GLOBAL_COUNTER = 1

def pseudonymize_name(name):
    """
    Replace a customer's real name with a consistent pseudonym.
    Example: "John Smith" -> "Customer_1"
    The same name always maps to the same pseudonym during a single script run.
    """
    global GLOBAL_COUNTER
    
    if pd.isnull(name) or not isinstance(name, str):
        return ""

    name = name.strip()
    if name not in NAME_MAP:
        # Create a new pseudonym
        NAME_MAP[name] = f"Customer_{GLOBAL_COUNTER}"
        GLOBAL_COUNTER += 1

    return NAME_MAP[name]

locale.setlocale(locale.LC_ALL, '')  # Use '' for system's default locale (e.g., USD for the US)

def process_file(file_path):
    """Reads an Excel file with a known structure (header=4),
    standardizes columns, and adds a 'day of week' column."""
    if not os.path.exists(file_path):
        print(f"Error: The file at path {file_path} does not exist.")
        # Return empty DataFrame with expected structure
        return pd.DataFrame(columns=[
            "order id", "order time", "budtender name", "customer name", "customer type",
            "vendor name", "product name", "category", "package id", "batch id",
            "external package id", "total inventory sold", "unit weight sold", "total weight sold",
            "gross sales", "inventory cost", "discounted amount", "loyalty as discount",
            "net sales", "return date", "upc gtin (canada)", "provincial sku (canada)",
            "producer", "order profit", "day of week"
        ])

    # OPTIONAL: capture pandas warnings and re-emit with file context
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        df = pd.read_excel(file_path, header=4)
        for w in caught:
            # Show the file this warning is associated with
            print(f"⚠️ [{os.path.abspath(file_path)}] {w.category.__name__}: {w.message}")
    df.columns = df.columns.str.strip().str.lower()

    # Standard set of columns for Dutchie sales exports
    df.columns = [
        "order id", "order time", "budtender name", "customer name", "customer type",
        "vendor name", "product name", "category", "package id", "batch id",
        "external package id", "total inventory sold", "unit weight sold", "total weight sold",
        "gross sales", "inventory cost", "discounted amount", "loyalty as discount",
        "net sales", "return date", "upc gtin (canada)", "provincial sku (canada)",
        "producer", "order profit"
    ]

    # Convert order time to datetime, then create day-of-week
    df['order time'] = pd.to_datetime(df['order time'], errors='coerce')
    df['day of week'] = df['order time'].dt.strftime('%A')
    df['customer name'] = df['customer name'].apply(pseudonymize_name)
    # NEW: tag rows with their source file and store code for later debug/traceability
    df['__source_file'] = os.path.basename(file_path)
    # Infer store code from filename like "salesMV.xlsx" -> "MV"
    _bn = os.path.basename(file_path)
    _m = re.search(r"sales([A-Za-z]+)\.xlsx$", _bn)
    df['__store'] = _m.group(1).upper() if _m else ""
    # Debug: show shape and columns
    # print(f"DEBUG: Successfully read {file_path}")
    # print(f"DEBUG: {file_path} shape: {df.shape}")
    # print(f"DEBUG: {file_path} columns: {list(df.columns)}")
    return df

import numpy as np
def apply_discounts_and_kickbacks(data, discount, kickback):
    """
    Adds discount/kickback columns and extra calculated metrics to the DataFrame.
    """

    # 1) Original discount/kickback
    data['discount amount'] = data['gross sales'] * discount
    data['kickback amount'] = data['inventory cost'] * kickback

    # 2) Net Profit = Gross Sales - Inventory Cost - Discount Amount
    data['net profit'] = data['gross sales'] - data['inventory cost'] - data['discount amount']

    # 3) Gross Margin % = ((Gross Sales - Inventory Cost) / Gross Sales) * 100
    #    Avoid division by zero using np.where
    data['gross margin %'] = np.where(
        data['gross sales'] != 0,
        (data['gross sales'] - data['inventory cost']) / data['gross sales'] * 100,
        0
    )

    # 4) Discount % = (Discount Amount / Gross Sales) * 100
    data['discount %'] = np.where(
        data['gross sales'] != 0,
        (data['discount amount'] / data['gross sales']) * 100,
        0
    )

    # 5) Profit Margin % = (Net Profit / Gross Sales) * 100
    data['profit margin %'] = np.where(
        data['gross sales'] != 0,
        (data['net profit'] / data['gross sales']) * 100,
        0
    )

    # 6) Break-Even Sales = Inventory Cost + Discount Amount
    data['break-even sales'] = data['inventory cost'] + data['discount amount']

    # 7) Efficiency Ratio = Gross Sales / Inventory Cost
    data['efficiency ratio'] = np.where(
        data['inventory cost'] != 0,
        data['gross sales'] / data['inventory cost'],
        0
    )

    # 8) Discount Impact % = (Discount Amount / Inventory Cost) * 100
    data['discount impact %'] = np.where(
        data['inventory cost'] != 0,
        (data['discount amount'] / data['inventory cost']) * 100,
        0
    )

    # 9) Sales to Cost Ratio = Gross Sales / Inventory Cost
    data['sales to cost ratio'] = np.where(
        data['inventory cost'] != 0,
        data['gross sales'] / data['inventory cost'],
        0
    )

    return data

brand_criteria = {
    'Hashish': {
        'vendors': ['Zenleaf LLC','Center Street Investments Inc.','Garden Of Weeden Inc.','BTC Ventures'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Hashish'] 
    },
    'Jeeter': {
        'vendors': ['Med For America Inc.'],
        'brands': ['Jeeter'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        # default (used by any rule that doesn't override discount/kickback)
        'discount': 0.50,
        'kickback': 0.30,

        # 'rules': [
        #     {
        #         'rule_name': 'Jeeter - Monday (50/30)',
        #         'days': ['Monday'],
        #         'discount': 0.50,
        #         'kickback': 0.30,
        #     },
        #     {
        #         'rule_name': 'Jeeter - Tue-Sun (40/20)',
        #         'days': ['Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
        #         # discount/kickback inherited from base (0.40 / 0.20)
        #     },
        # ],

        # optional: only if you want to restrict stores
        # 'stores': ['MV','LM','SV','LG','NC','WP'],
    },  
    'Kiva': {
        'vendors': ['KIVA / LCISM CORP', 'Vino & Cigarro, LLC','Garden Of Weeden Inc.','PuffCo'],
        'days': ['Monday','Wednesday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Terra', 'Petra', 'KIVA', 'Lost Farms', 'Camino']
    },
    'Dabwoods': {
        'vendors': ['The Clear Group Inc.','Decoi','Garden Of Weeden Inc.','Garden Of Weeden'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'categories': ['Disposables','Cartridges'],
        'excluded_phrases': ['DabBar X','Cart'],
        'brands': ['Dabwoods','DabBar']
    },
     'Time Machine': {
         'vendors': ['Vino & Cigarro, LLC','Garden Of Weeden Inc.','KIVA / LCISM CORP','Garden Of Weeden'],
         'days': ['Tuesday','Friday'],
         'discount': 0.50,
         'kickback': 0.25,
         'brands': ['Time Machine']
     },
     'Pacific Stone': {
         'vendors': ['Vino & Cigarro, LLC','KIVA / LCISM CORP', 'Garden Of Weeden Inc.','Pacific Stone','Garden Of Weeden'],
         'days': ['Monday','Thursday'],
         'discount': 0.50,
         'kickback': 0.25,
         'brands': ['Pacific Stone']
    },
    'Heavy Hitters': {
        'vendors': ['Fluids Manufacturing Inc.','Garden Of Weeden Inc.'],
        'days': ['Friday','Saturday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Heavy Hitters']
    },
    'Almora': {
        'vendors': ['Fluids Manufacturing Inc.','Garden Of Weeden Inc.','Vino & Cigarro, LLC'],
        'days': ['Sunday','Saturday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Almora']
    },
    'WYLD/GoodTide': {
        'vendors': ['2020 Long Beach LLC'],
        'days': ['Friday','Saturday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Wyld', 'Good Tide']
    },
    'Jetty': {
        'vendors': ['KIVA / LCISM CORP', 'Vino & Cigarro, LLC','Garden Of Weeden Inc.','Garden Of Weeden'],
        'days': ['Thursday'],
        'discount': 0.50,
        'kickback': 0.25,
        #'excluded_phrases': ['Jetty | Cart 1g |'],
        #'include_phrases': ['SVL','ULR',],
        'brands': ['Jetty']
    },
    'Preferred': {
        'vendors': ['Garden Of Weeden Inc.','Helios | Hypeereon Corporation','Garden Of Weeden'],
        'days': ['Monday','Wednesday'],
        'discount': 0.50,
        'kickback': 0.25,
        'brands': ['Preferred Gardens']
    },
    'Kikoko': {
        'vendors': ['Garden Of Weeden Inc.'],
        'days': ['Wednesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Kikoko']
    },
    'TreeSap': {
        'vendors': ['Zenleaf LLC','Center Street Investments Inc.','Fluids Manufacturing Inc.','Garden Of Weeden Inc.'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.25, 
        'brands': ['TreeSap']
    },
      'Made': { 
        'vendors': ['Garden Of Weeden Inc.','Garden Of Weeden'],
        'days': ['Friday','Saturday'],
        'discount': 0.50,
        'kickback': 0.30,
        'categories': ['Pre-Rolls','Flower','Eighths'], 
        'brands': ['Made |']
    }, 
    
      'Made-Eddys': { 
        'vendors': ['Garden Of Weeden Inc.','Garden Of Weeden'],
        'days': ['Friday','Saturday'],
        'discount': 0.50,
        'kickback': 0.30,
        'categories': ['Edibles'], 
        'brands': ['Made |']
    }, 
      'Yada Yada': { 
        'vendors': ['Fluids Manufacturing Inc.'],
        'days': ['Wednesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Yada Yada |']
    }, 
      'Eureka': { 
        'vendors': ['Light Box Leasing Corp.'],
        'days': ['Monday','Tuesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Eureka |']
    }, 
    'Ember Valley': { 
        'vendors': ['LB Atlantis LLC', 'Garden Of Weeden Inc.', 'Courtney Lang', 'Hilife Group MV , LLC', 'Ember Valley', 'Helios | Hypeereon Corporation'],
        'days': ['Thursday','Tuesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Ember Valley |','EV |']
    }, 
    'Cake': { 
        'vendors': ['ThirtyOne Labs, LLC'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.40,
        'kickback': 0.0,
        'brands': ['Cake |']
    }, 
    # 'Green Dawg': { 
    #     'vendors': ['Artisan Canna Cigars LLC'],
    #     'days': ['Thursday'],
    #     'discount': 0.50,
    #     'kickback': 0.30,
    #     #'categories': [''], 
    #     'brands': ['Green Dawg |']
    # }, 
    'Mary Medical': { 
        'vendors': ["Mary's Tech CA, Inc.",'BRB California LLC', 'Garden Of Weeden Inc.', 'Broadway Alliance, LLC','Garden Of Weeden','Garden Society / LCISM Corp'],
        'days': ['Thursday'],
        'discount': 0.50,
        'kickback': 0.30,
        #'categories': [''], 
        'brands': ["Mary's Medicinals |"]
    }, 
    'LA FARMS': { 
        'vendors': ["LA Family Farms LLC",'Los Angeles Family Farms LLC'],
        'days': ['Friday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        #'categories': [''], 
        'brands': ["L.A.FF |"]
    },  
    'COTC': { 
        'vendors': ["TERPX COTC/WCTC (Riverside)"],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.40,
        'kickback': 0.0,
        'brands': ["COTC |"]
    },  
    # 'Master Makers': {
    #     'vendors': ['Broadway Alliance, LLC'],
    #     'days': ['Tuesday','Thursday'],
    #     'discount': 0.50,
    #     'kickback': 0.30,
    #     'brands': ['Master Makers |']
    # }, 
    'Dixie': {
        'vendors': ['Broadway Alliance, LLC','BRB California LLC', 'Garden Of Weeden Inc.','Hilife Group MV , LLC','Garden Of Weeden',"Mary's Tech CA, Inc."],
        'days': ['Saturday','Thursday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Dixie']
    },
    "710": {
        'vendors': ['Fluids Manufacturing Inc.'],
        'days': ['Monday'],
        'discount': 0.50, #9/31 last day
        'kickback': 0.30,
        'brands': ['710']
    },
    "Blem": {
        'vendors': ['SSAL HORTICULTURE LLC'],
        'days': ['Thursday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['BLEM']
    },
       "P&B": {
        'vendors': ['Fluids Manufacturing Inc.'],
        'days': ['Sunday','Tuesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['P&B |']
    }, 
    "Drops": {
        'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.'],
        'days': ['Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Drops | ']
    },
    "SeedJunky": {
        'vendors': ['Seed Junky | LCISM Corp','Garden Of Weeden Inc.','Vino & Cigarro, LLC','Garden Of Weeden'],
        'days': ['Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Seed Junky']
    },
    "KEEF": {
        'vendors': ['Vino & Cigarro, LLC','Garden Of Weeden Inc.','KIVA / LCISM CORP','GB2, LLC'],
        'days': ['Tuesday','Wednesday'],
        'discount': 0.50,
        'kickback': 0.35,
        'brands': ['Keef']
    },
    "PlugnPlay": {
        'vendors': ['Vino & Cigarro, LLC','Garden Of Weeden Inc.','KIVA / LCISM CORP','IE Licensing, LLC'],
        'days': ['Monday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Plug n Play |','Plug N Play |']
    },
    "Sluggers": {
        'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.'],
        'brands': ['Sluggers'],
        'rules': [
            {
                'rule_name': 'Sluggers - Monday Sunday (50/30)',
                'days': ['Monday','Sunday'],
                'discount': 0.50,
                'kickback': 0.30,
            },
            {
                'rule_name': 'Sluggers - Tue-Sat (30/0)',
                'days': ['Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
                'discount': 0.30,
                'kickback': 0.0,
            },
        ],
    },
    "Turn": {
        'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.','Hilife Group MV , LLC'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Turn |']
    },
  "Level": {
    "vendors": ["Garden Of Weeden Inc.", "Vino & Cigarro, LLC"],
    "brands": ["Level |", "LEVEL |"],

    "rules": [
      { # Level (NOT 10mg)
        "rule_name": "Level (non-10mg)",
        "days": ["Monday", "Tuesday"],
        "discount": 0.50,
        "kickback": 0.25,
        "excluded_phrases": ["10mg"],
      },
      { # Level 10mg
        "rule_name": "Level 10mg",
        "days": ["Monday", "Tuesday"],
        "discount": 0.50,
        "kickback": 0.30,
        "include_phrases": ["10mg"],
      },
    ],
  },
    "Raw Garden": {
        'vendors': ['Garden Of Weeden Inc.','Garden Of Weeden'],
        'days': ['Wednsday','Tuesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Raw Garden |']
    }, 
    "Claybourne": {
        'vendors': ['CI Distribution','Garden Of Weeden Inc.'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Claybourne |']
    }, 
    "Smokiez": {
        'vendors': ['Garden Of Weeden Inc.'],
        'days': ['Tuesday','Wednesday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['Smokies |']
    },
    # "Uncle Arnies": {
    #     'vendors': ['Garden Of Weeden Inc.','Vino & Cigarro, LLC','KIVA / LCISM CORP'],
    #     'days': ['Monday','Friday'],
    #     'discount': 0.50,
    #     'kickback': 0.30,
    #     'brands': ["Uncle Arnie's |"]
    # },
    "KANHA": {
        'vendors': ['Garden Of Weeden Inc.','Sunderstorm Bay LLC.'],
        'days': ['Thursday','Sunday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ["KANHA |",'Kanha | ']
    },
     "Kushy Punch": {
        'vendors': ['Garden Of Weeden Inc.','Varavo'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.0,
        'brands': ['Kushy Punch |']
    },
    # "Royal Blunts": {
    #     'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.','Royal M&D LLC','Bud Technology'],
    #     'days': ['Monday','Wednesday'],
    #     'discount': 0.50,
    #     'kickback': 0.30,
    #     'brands': ['Royal Blunts']
    # },
    # "Heady Heads": {
    #     'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.','HD Health Industries'],
    #     'days': ['Wednesday'],
    #     'discount': 0.50,
    #     'kickback': 0.30,
    #     'brands': ['Heady Heads |']
    # }, 
    "American Weed": {
        'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.'],
        'days': ['Wednesday','Thursday'],
        'discount': 0.50,
        'kickback': 0.30,
        'brands': ['American Weed']
    },    
    'Josh Wax': {  # merged
        'vendors': ['Zasp', 'Garden Of Weeden Inc.', 'Garden Of Weeden'],
        'brands': ['Josh Wax'],
        'discount': 0.40,
        'kickback': 0.00,

        'rules': [
            {
                'rule_name': 'Josh Wax - Friday (50/12)',
                'days': ['Friday'],
                'discount': 0.50,
                'kickback': 0.12,
            },
            {
                'rule_name': 'Josh Wax - Other days (40/0)',
                'days': ['Monday','Tuesday','Wednesday','Thursday','Saturday','Sunday'],
                # inherits base 0.40 / 0.00
            },
        ],
    },
    'Cam': {  # OFF INVOICE (merged)
        'vendors': [
            "California Artisanal Medicine (CAM)",
            "NC INVESTMENT GROUP, LLC",
            "Garden Of Weeden Inc.",
            "Garden Of Weeden",
        ],
        'brands': ['CAM |', 'CAM|', 'CAM '],  # safer than "CAM"
        'discount': 0.40,
        'kickback': 0.00,

        'rules': [
            {
                'rule_name': 'Cam - Saturday (50/12)',
                'days': ['Saturday'],
                'discount': 0.50,
                'kickback': 0.12,
            },
            {
                'rule_name': 'Cam - Sun-Fri (40/0)',
                'days': ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday'],
                # inherits base 0.40 / 0.00
            },
        ],
    },
    "PBR-NYF-STIDES": {
        'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.'],
        'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
        'discount': 0.50,
        'kickback': 0.0,
        'brands': ['NYF |','PBR |','St. Ides |'],
    },  
    'CLSICS': {'vendors': ['KIVA / LCISM CORP','Garden Of Weeden Inc.','Vino & Cigarro, LLC'],
            'days': ['Tuesday'],
            'discount': 0.5,
            'kickback': 0.3,
            'brands': ['CLSICS']},
    'Nasha': {'vendors': ['KIVA / LCISM CORP', 'Vino & Cigarro, LLC'],
           'days': ['Tuesday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Nasha']},
    'Ball Family Farms': {'vendors': ['Ball Family Farms Corporation'],
           'days': ['Saturday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Ball Family Farms']},
    'Just J': {'vendors': ['Rukli Inc (Tyson)'],
           'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['JJ Dragon']},
    'Sol Flora': {'vendors': ['Twisted Roots, Inc.'],
           'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
           'discount': 0.5,
           'kickback': 0.0,
           'brands': ['Sol Flora']},
    'Planta': {'vendors': ['Higher Logic LLC','Garden Of Weeden Inc.'],
           'days': ['Monday','Sunday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Planta |']},
    'Seed Junky': {'vendors': ['Higher Logic LLC','Garden Of Weeden Inc.','Garden Of Weeden','Seed Junky | LCISM Corp'],
           'days': ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'],
           'categories':['Eighths'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Seed Junky |']},
    'Sauce': {'vendors': ['Garden Of Weeden','Garden Of Weeden Inc.'],
           'days': ['Friday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Sauce |']},
    'Cannabiotix (CBX)': {'vendors': ['Four Star Distribution and Delivery LLC','Highstar Distribution LLC','Hilife LM'],
           'days': ['Tuesday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['CBX |']},
    'Heirbloom': {'vendors': ['Four Star Distribution and Delivery LLC','Highstar Distribution LLC','Hilife LM'],
           'days': ['Wednesday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['HB |']},
    'Happy Fruit': {'vendors': ['Garden Of Weeden Inc.','Garden Of Weeden'],
           'days': ['Tuesday','Thursday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Happy Fruit |']},
    'Highatus': {'vendors': ['Four Star Distribution and Delivery LLC','Highstar Distribution LLC','Hilife LM'],
           'days': ['Tuesday'],
           'discount': 0.5,
           'kickback': 0.3,
           'brands': ['Highatus |']},
}

def style_summary_sheet(sheet, brand_name):
    """
    Styles the Summary sheet:
      - A bold title in row 1
      - Headers in row 2 (gray background, centered)
      - Data starts in row 3
      - Freeze pane at A3
      - Banded row styling for data
      - Currency/date formatting as needed
    """
    max_col = sheet.max_column
    max_row = sheet.max_row

    # 1) Big title in row 1
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"{brand_name.upper()} SUMMARY REPORT"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # 2) Style header row (row 2)
    header_row_idx = 2
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col_idx in range(1, max_col + 1):
        cell = sheet.cell(row=header_row_idx, column=col_idx)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        cell.border = thin_border

    # 3) Freeze panes at row 3
    sheet.freeze_panes = "A3"

    # 4) Style data rows (row 3 downward)
    for row_idx in range(3, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            hdr_val = sheet.cell(row=header_row_idx, column=col_idx).value
            lower_hdr = str(hdr_val).lower() if hdr_val is not None else ""
            if "owed" in lower_hdr:
                # Format as currency
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif "gross sales" in lower_hdr or "discount amount" in lower_hdr:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            elif hdr_val and ("date" in str(hdr_val).lower()):
                # Format as date
                cell.number_format = "YYYY-MM-DD"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Banded row coloring
            if row_idx % 2 == 1:  # Odd data row
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 5) Auto-fit column widths
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_length = len(str(val))
                if val_length > max_length:
                    max_length = val_length
        sheet.column_dimensions[col_letter].width = max_length + 2

def style_worksheet(sheet):
    """
    Similar styling for other sheets like MV_Sales, LM_Sales, SV_Sales, etc.
    """
    max_col = sheet.max_column
    # Make header row bold and center-aligned
    for col_idx in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit column width
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_length = len(str(val))
                if val_length > max_length:
                    max_length = val_length
        sheet.column_dimensions[column_letter].width = max_length + 2

    # Freeze row 1
    sheet.freeze_panes = "A2"

def style_top_sellers_sheet(sheet):
    """
    Styles a 'Top Sellers' sheet:
      - Bold header
      - Currency formatting for "Gross Sales"
      - Alternating row colors
      - Auto-fit columns
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    max_col = sheet.max_column

    # Header row
    for col_idx in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Data rows
    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            # "Gross Sales" is column 2 in "Top Sellers"
            if col_idx == 2:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

            # Alternating row color
            if row_idx % 2 == 1:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Auto-fit columns
    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_length = len(str(val))
                if val_length > max_length:
                    max_length = val_length
        sheet.column_dimensions[column_letter].width = max_length + 2
def discount_for_store(base_discount: float, store_code: str) -> float:
            """
            Returns the effective discount for a given store.
            Store 'WP' has special rules: 0.5 → 0.3, 0.4 → 0.2.
            """
            if store_code == 'WP':
                if base_discount == 0.50:
                    return 0.50
                elif base_discount == 0.40:
                    return 0.40
            return base_discount
from collections import defaultdict

def print_unknown_vendors(brand: str, criteria: dict, dataframes: list):
    """
    For the given brand and its criteria, scan all provided dataframes and
    print unknown vendors along with the Excel file(s) they came from.
    """
    brand_keywords = set(criteria.get('brands', []))
    expected_vendors = set(criteria.get('vendors', []))
    if not brand_keywords:
        return

    def _matches_brand(name: str) -> bool:
        s = str(name or "")
        s_low = s.lower()
        return any(b.lower() in s_low for b in brand_keywords)

    unknown_map = defaultdict(set)  # vendor -> set of source files
    days = set(criteria.get('days', []))

    for df in dataframes:
        if df is None or df.empty:
            continue
        if 'day of week' not in df.columns or 'product name' not in df.columns or 'vendor name' not in df.columns:
            continue
        # Day filter first
        day_df = df[df['day of week'].isin(days)]
        if day_df.empty:
            continue
        # Then brand match
        matched = day_df[day_df['product name'].apply(_matches_brand)]
        if matched.empty:
            continue
        # Collect unknown vendors with their source files
        for _, row in matched.iterrows():
            vendor = str(row.get('vendor name', '')).strip()
            if not vendor or vendor in expected_vendors:
                continue
            src = row.get('__source_file', '<unknown file>')
            unknown_map[vendor].add(src)

    if unknown_map:
        print(f"\n⚠️⚠️⚠️ Brand '{brand}' has unknown vendor(s):")
        for v, files in sorted(unknown_map.items()):
            files_list = ", ".join(sorted(files))
            print(f"   - {v}: {files_list}")
        print(f"👉 Consider adding to brand_criteria['{brand}']['vendors']\n")
DEFAULT_STORES = ["MV", "LM", "SV", "LG", "NC", "WP"]
DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
ALL_DAYS_SET = set(DAY_ORDER)

def normalize_rules(criteria):
    """
    Supports:
      - {"vendors":..., "days":..., ...}                         # single rule
      - {"vendors":..., "brands":..., "rules":[{...},{...}]}     # base + rules overrides
      - [{...},{...}]                                           # list of rules (no base)
    Returns: list[effective_rule_dict]
    """
    if isinstance(criteria, list):
        base = {}
        rules = criteria
    else:
        base = dict(criteria or {})
        rules = base.pop("rules", None) or [{}]

    out = []
    for i, r in enumerate(rules, 1):
        effective = dict(base)
        effective.update(r or {})

        effective.setdefault("rule_name", f"Rule {i}")
        effective.setdefault("stores", base.get("stores", DEFAULT_STORES))

        # Safety: keep the old “must filter by vendors+days” behavior unless you explicitly provide otherwise
        if "vendors" not in effective:
            effective["vendors"] = base.get("vendors", [])
        if "days" not in effective:
            effective["days"] = base.get("days", [])

        out.append(effective)
    return out

def _contains_any(haystack_series, needles):
    needles = [str(n).lower() for n in (needles or []) if str(n).strip()]
    if not needles:
        return haystack_series.notna()  # no-op
    s = haystack_series.astype(str).str.lower()
    return s.apply(lambda x: any(n in x for n in needles))

def filter_by_rule(df, rule):
    """
    Apply all filters for a single rule.
    """
    out = df

    # stores filter happens outside (because df already store-specific)

    vendors = rule.get("vendors") or []
    days = rule.get("days") or []
    categories = rule.get("categories") or []
    brands = rule.get("brands") or []
    include_phrases = rule.get("include_phrases") or []
    excluded_phrases = rule.get("excluded_phrases") or []

    if vendors:
        out = out[out["vendor name"].isin(vendors)]
    if days:
        out = out[out["day of week"].isin(days)]
    if categories:
        out = out[out["category"].isin(categories)]

    if brands:
        out = out[_contains_any(out["product name"], brands)]

    if include_phrases:
        regex = "|".join(re.escape(p) for p in include_phrases)
        out = out[out["product name"].astype(str).str.contains(regex, case=False, na=False, regex=True)]

    if excluded_phrases:
        regex = "|".join(re.escape(p) for p in excluded_phrases)
        out = out[~out["product name"].astype(str).str.contains(regex, case=False, na=False, regex=True)]

    return out

def days_text_from_rules(rules):
    days = set()
    for r in rules:
        days.update(r.get("days", []) or [])
    if days == ALL_DAYS_SET:
        return "Everyday"
    return ", ".join([d for d in DAY_ORDER if d in days])

def build_brand_store_data(brand, criteria, store_data):
    """
    For a single brand:
      - Apply ALL rules per store
      - Prevent double-counting
      - Collect:
          1) Per-store combined data
          2) Per-rule combined data (NEW)
    """
    rules = normalize_rules(criteria)

    remaining = {
        code: df.copy()
        for code, df in store_data.items()
        if df is not None and not df.empty
    }

    collected_by_store = {code: [] for code in store_data.keys()}
    collected_by_rule = {}  # <-- NEW

    for rule in rules:
        rule_name = rule.get("rule_name", "Unnamed Rule")
        collected_by_rule[rule_name] = []

        allowed_stores = set(rule.get("stores", DEFAULT_STORES))

        for store_code, df in list(remaining.items()):
            if store_code not in allowed_stores:
                continue

            matched = filter_by_rule(df, rule)
            if matched.empty:
                continue

            matched = matched.copy()
            matched["__deal_rule"] = rule_name
            matched["__store"] = store_code

            # Apply per-rule math
            if {"gross sales", "inventory cost"}.issubset(matched.columns):
                d = float(rule.get("discount", 0.0))
                k = float(rule.get("kickback", 0.0))
                d = discount_for_store(d, store_code)
                matched = apply_discounts_and_kickbacks(matched, d, k)

            collected_by_store[store_code].append(matched)
            collected_by_rule[rule_name].append(matched)

            # Prevent double-counting
            remaining[store_code] = df.drop(index=matched.index, errors="ignore")

    # Combine store outputs
    store_out = {
        store: (
            pd.concat(frames, ignore_index=True)
            if frames else pd.DataFrame()
        )
        for store, frames in collected_by_store.items()
    }

    # Combine rule outputs
    rule_out = {
        rule: (
            pd.concat(frames, ignore_index=True)
            if frames else pd.DataFrame()
        )
        for rule, frames in collected_by_rule.items()
    }

    return store_out, rule_out, rules
def build_rule_summary(rule_df, rule_name, brand, start_date, end_date, days_text):
    """
    Builds a Summary-style dataframe for a single rule.
    """
    if rule_df.empty:
        return pd.DataFrame()

    summary = (
        rule_df
        .groupby("__store", as_index=False)
        .agg({
            "gross sales": "sum",
            "inventory cost": "sum",
            "discount amount": "sum",
            "kickback amount": "sum",
        })
    )

    summary.rename(
        columns={
            "__store": "Store",
            "kickback amount": "Kickback Owed",
        },
        inplace=True
    )

    summary["Days Active"] = days_text
    summary["Date Range"] = f"{start_date} to {end_date}"
    summary["Brand"] = brand
    summary["Rule"] = rule_name
    summary["Margin"] = None

    # Match Summary column order
    summary = summary[
        [
            "Store",
            "Kickback Owed",
            "Days Active",
            "Date Range",
            "gross sales",
            "inventory cost",
            "discount amount",
            "Margin",
            "Brand",
            "Rule",
        ]
    ]

    return summary

def run_deals_reports():
    """
    Multi-rule version:
      - Each brand can have one rule OR many rules (criteria["rules"] list)
      - All matched rows across rules are combined into ONE report per brand
      - Discount/kickback is applied PER ROW using the rule that matched it
      - Prevents double-counting when rules overlap (earlier rule wins)
    """
    output_dir = "brand_reports"
    old_dir = "old"

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    Path(old_dir).mkdir(parents=True, exist_ok=True)

    # Archive old reports before generating new ones
    for file in os.listdir(output_dir):
        full_path = os.path.join(output_dir, file)
        if file.endswith(".xlsx") and os.path.isfile(full_path):
            dest_path = os.path.join(old_dir, file)
            if os.path.exists(dest_path):
                os.remove(dest_path)
            shutil.move(full_path, dest_path)

    # Read store files (process_file already returns empty DF if missing)
    mv_data = process_file("files/salesMV.xlsx")
    lm_data = process_file("files/salesLM.xlsx")
    sv_data = process_file("files/salesSV.xlsx")
    lg_data = process_file("files/salesLG.xlsx")
    nc_data = process_file("files/salesNC.xlsx")
    wp_data = process_file("files/salesWP.xlsx")

    store_data = {
        "MV": mv_data, "LM": lm_data, "SV": sv_data,
        "LG": lg_data, "NC": nc_data, "WP": wp_data
    }

    consolidated_summary = []
    results_for_app = []

    for brand, criteria in brand_criteria.items():
        if not isinstance(criteria, (dict, list)):
            print(f"[SKIP] Brand '{brand}' has invalid criteria type. Skipping.")
            continue

        # Optional: vendor sanity check across ALL rules (union of vendors/brands/days)
        try:
            _rules_for_debug = normalize_rules(criteria)
            _dbg_vendors = set()
            _dbg_brands = set()
            _dbg_days = set()
            for r in _rules_for_debug:
                _dbg_vendors.update(r.get("vendors", []) or [])
                _dbg_brands.update(r.get("brands", []) or [])
                _dbg_days.update(r.get("days", []) or [])
            if _dbg_vendors and _dbg_brands and _dbg_days:
                print_unknown_vendors(
                    brand,
                    {"vendors": list(_dbg_vendors), "brands": list(_dbg_brands), "days": list(_dbg_days)},
                    [mv_data, lm_data, sv_data, lg_data, nc_data, wp_data],
                )
        except Exception:
            pass

        # --- NEW: apply multiple rules per brand and combine into ONE report ---
        brand_store_data, rule_data, rules = build_brand_store_data(brand, criteria, store_data)


        mv_brand_data = brand_store_data["MV"]
        lm_brand_data = brand_store_data["LM"]
        sv_brand_data = brand_store_data["SV"]
        lg_brand_data = brand_store_data["LG"]
        nc_brand_data = brand_store_data["NC"]
        wp_brand_data = brand_store_data["WP"]

        print(
            f"DEBUG: {brand} - After rule filtering => "
            f"MV: {mv_brand_data.shape}, LM: {lm_brand_data.shape}, "
            f"SV: {sv_brand_data.shape}, LG: {lg_brand_data.shape}, "
            f"NC: {nc_brand_data.shape}, WP: {wp_brand_data.shape}"
        )

        if (
            mv_brand_data.empty and lm_brand_data.empty and sv_brand_data.empty and
            lg_brand_data.empty and nc_brand_data.empty and wp_brand_data.empty
        ):
            print(f"DEBUG: No data remains for brand '{brand}'. Skipping.")
            continue

        # ---- Date range across all stores used for this brand ----
        store_dfs = [mv_brand_data, lm_brand_data, sv_brand_data, lg_brand_data, nc_brand_data, wp_brand_data]
        possible_starts = [
            df["order time"].min()
            for df in store_dfs
            if (df is not None and not df.empty and "order time" in df.columns)
        ]
        possible_ends = [
            df["order time"].max()
            for df in store_dfs
            if (df is not None and not df.empty and "order time" in df.columns)
        ]

        if not possible_starts or not possible_ends:
            print(f"DEBUG: Brand '{brand}' had data, but no valid date range. Skipping.")
            continue

        start_date = min(possible_starts).strftime("%Y-%m-%d")
        end_date = max(possible_ends).strftime("%Y-%m-%d")
        date_range = f"{start_date}_to_{end_date}"

        # ---- Summary rows per store ----
        def build_summary(df, store_name, include_units=False):
            if df is None or df.empty:
                return pd.DataFrame(
                    columns=["gross sales", "inventory cost", "discount amount", "kickback amount", "location"]
                )

            agg_map = {
                "gross sales": "sum",
                "inventory cost": "sum",
                "discount amount": "sum",
                "kickback amount": "sum",
            }
            if include_units and "total inventory sold" in df.columns:
                agg_map["total inventory sold"] = "sum"

            summary = df.agg(agg_map).to_frame().T
            summary["location"] = store_name
            return summary

        if isinstance(criteria, dict):
            want_units = bool(criteria.get("include_units", False))
        else:
            want_units = any(bool(r.get("include_units", False)) for r in rules)

        mv_summary = build_summary(mv_brand_data, "Mission Valley", include_units=want_units)
        lm_summary = build_summary(lm_brand_data, "La Mesa", include_units=want_units)
        sv_summary = build_summary(sv_brand_data, "Sorrento Valley", include_units=want_units)
        lg_summary = build_summary(lg_brand_data, "Lemon Grove", include_units=want_units)
        nc_summary = build_summary(nc_brand_data, "National City", include_units=want_units)
        wp_summary = build_summary(wp_brand_data, "Wildomar Palomar", include_units=want_units)

        brand_summary = pd.concat(
            [mv_summary, lm_summary, sv_summary, lg_summary, nc_summary, wp_summary],
            ignore_index=True,
        )

        # Days Active now comes from ALL rules (union)
        days_text = days_text_from_rules(rules)

        # Rename + add metadata
        brand_summary.rename(
            columns={
                "location": "Store",
                "kickback amount": "Kickback Owed",
                "total inventory sold": "Units Sold",
            },
            inplace=True,
        )
        brand_summary["Days Active"] = days_text
        brand_summary["Date Range"] = f"{start_date} to {end_date}"
        brand_summary["Brand"] = brand
        brand_summary["Margin"] = None

        col_order = [
            "Store", "Kickback Owed", "Days Active", "Date Range",
            "gross sales", "inventory cost", "discount amount", "Margin", "Brand",
        ]
        if want_units:
            col_order.append("Units Sold")

        brand_summary = brand_summary[[c for c in col_order if c in brand_summary.columns]]

        consolidated_summary.append(brand_summary)

        # ---- Create brand-level Excel ----
        safe_brand_name = brand.replace("/", " ")
        output_filename = os.path.join(output_dir, f"{safe_brand_name}_report_{date_range}.xlsx")
        print(f"DEBUG: Creating {output_filename} for brand '{brand}'...")

        combined_df = pd.concat(store_dfs, ignore_index=True)
        if not combined_df.empty and "gross sales" in combined_df.columns:
            top_sellers_df = (
                combined_df.groupby("product name", as_index=False)
                .agg({"gross sales": "sum"})
                .sort_values(by="gross sales", ascending=False)
                .head(20)
            )
            top_sellers_df.rename(
                columns={"product name": "Product Name", "gross sales": "Gross Sales"},
                inplace=True,
            )
        else:
            top_sellers_df = pd.DataFrame(columns=["Product Name", "Gross Sales"])

        with pd.ExcelWriter(output_filename) as writer:
            brand_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=1)

            if not mv_brand_data.empty:
                mv_brand_data.to_excel(writer, sheet_name="MV_Sales", index=False)
            if not lm_brand_data.empty:
                lm_brand_data.to_excel(writer, sheet_name="LM_Sales", index=False)
            if not sv_brand_data.empty:
                sv_brand_data.to_excel(writer, sheet_name="SV_Sales", index=False)
            if not lg_brand_data.empty:
                lg_brand_data.to_excel(writer, sheet_name="LG_Sales", index=False)
            if not nc_brand_data.empty:
                nc_brand_data.to_excel(writer, sheet_name="NC_Sales", index=False)
            if not wp_brand_data.empty:
                wp_brand_data.to_excel(writer, sheet_name="WP_Sales", index=False)

            top_sellers_df.to_excel(writer, sheet_name="Top Sellers", index=False)
            # --- NEW: Rule-level sheets ---
            for rule_name, rule_df in rule_data.items():
                if rule_df is None or rule_df.empty:
                    continue

                safe_rule_name = (
                    rule_name
                    .replace("/", " ")
                    .replace("(", "")
                    .replace(")", "")
                    .replace("%", "")
                )

                sheet_name = f"Rule - {safe_rule_name}"
                sheet_name = sheet_name[:31]  # Excel limit

                for rule_name, rule_df in rule_data.items():
                    if rule_df is None or rule_df.empty:
                        continue

                    safe_rule_name = (
                        rule_name
                        .replace("/", " ")
                        .replace("(", "")
                        .replace(")", "")
                        .replace("%", "")
                    )

                    sheet_name = f"Rule - {safe_rule_name}"[:31]

                    rule_summary_df = build_rule_summary(
                        rule_df=rule_df,
                        rule_name=rule_name,
                        brand=brand,
                        start_date=start_date,
                        end_date=end_date,
                        days_text=days_text,
                    )

                    rule_summary_df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                        startrow=1
                    )
        # Inject Margin formulas (same layout as before: Margin is column H)
        wb = load_workbook(output_filename)
        summary_sheet = wb["Summary"]

        data_start_row = 3
        data_end_row = data_start_row + brand_summary.shape[0] - 1

        for row_idx in range(data_start_row, data_end_row + 1):
            summary_sheet.cell(row=row_idx, column=8).value = (
                f"=((E{row_idx}-G{row_idx})-(F{row_idx}-B{row_idx}))/(E{row_idx}-G{row_idx})"
            )

        wb.save(output_filename)

        # Style sheets
        wb = load_workbook(output_filename)
        if "Summary" in wb.sheetnames:
            style_summary_sheet(wb["Summary"], brand)
        for s in ["MV_Sales", "LM_Sales", "SV_Sales", "LG_Sales", "NC_Sales", "WP_Sales"]:
            if s in wb.sheetnames:
                style_worksheet(wb[s])
        if "Top Sellers" in wb.sheetnames:
            style_top_sellers_sheet(wb["Top Sellers"])
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("Rule -"):
                style_worksheet(wb[sheet_name])

        wb.save(output_filename)

        total_owed = float(pd.to_numeric(brand_summary.get("Kickback Owed"), errors="coerce").fillna(0).sum())
        results_for_app.append({"brand": brand, "owed": total_owed, "start": start_date, "end": end_date})

    # ---- Consolidated Summary ----
    if consolidated_summary:
        final_df = pd.concat(consolidated_summary, ignore_index=True)

        if results_for_app:
            all_starts = [r["start"] for r in results_for_app if r.get("start")]
            all_ends = [r["end"] for r in results_for_app if r.get("end")]
            overall_start = min(all_starts) if all_starts else start_date
            overall_end = max(all_ends) if all_ends else end_date
            overall_range = f"{overall_start}_to_{overall_end}"
        else:
            overall_range = date_range

        consolidated_file = os.path.join(output_dir, f"consolidated_brand_report_{overall_range}.xlsx")
        print(f"DEBUG: Creating consolidated summary => {consolidated_file}")

        with pd.ExcelWriter(consolidated_file) as writer:
            final_df.to_excel(writer, sheet_name="Consolidated_Summary", index=False, startrow=1)

        wb = load_workbook(consolidated_file)
        if "Consolidated_Summary" in wb.sheetnames:
            sheet = wb["Consolidated_Summary"]
            data_start_row = 3
            data_end_row = data_start_row + final_df.shape[0] - 1

            for row_idx in range(data_start_row, data_end_row + 1):
                sheet.cell(row=row_idx, column=8).value = (
                    f"=((E{row_idx}-G{row_idx})-(F{row_idx}-B{row_idx}))/(E{row_idx}-G{row_idx})"
                )

            style_summary_sheet(sheet, "ALL_BRANDS")

        wb.save(consolidated_file)
        print("Individual brand reports + consolidated report have been saved.")
    else:
        print("No brand data found; no Excel files generated.")

    return results_for_app

if __name__ == "__main__":
    data = run_deals_reports()
    print("Results for app:", data)