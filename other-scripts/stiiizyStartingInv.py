#!/usr/bin/env python3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────── Config ─────────────────────────
FILES_DIR = "files"
OUTPUT_DIR = "doneReports"
TARGET_BRAND = "Stiiizy"

# Categories that get 30% kickback; all others get 15%
SPECIAL_CATEGORIES = ["Disposables", "Cartridges", "Gummies", "Edibles", "Accessories"]

# Columns to keep in the final report (order matters)
KEEP_COLUMNS = ["available", "product", "category", "cost", "Total Cost", "Kickback Owed"]


# Summary placement:
SUMMARY_ROWS_USED = 4     # rows 1–4 (Title + 3 totals)
SPACER_ROWS = 1           # exactly one blank row between summary and table
HEADER_ROW = SUMMARY_ROWS_USED + SPACER_ROWS + 1  # -> 6 (header at row 6)
# first data row will be HEADER_ROW + 1
FREEZE_AT_ROW = HEADER_ROW + 1  # freeze at first row under header (sticky header)

# ───────────────────────── Core Logic ─────────────────────────
def process_file(file_path: Path):
    """Read CSV, filter Stiiizy, calculate kickbacks, and save to Excel with a tidy report."""
    df = pd.read_csv(file_path)

    # Normalize column names
    df.columns = df.columns.str.strip().str.lower()

    # Guard rails
    required = {"brand", "category", "cost"}
    if not required.issubset(set(df.columns)):
        print(f"⚠️ {file_path.name} missing required columns {required - set(df.columns)}. Skipping.")
        return

    # Filter by brand (case-insensitive, whitespace-safe)
    df = df[df["brand"].astype(str).str.strip().str.lower() == TARGET_BRAND.lower()].copy()
    if df.empty:
        print(f"ℹ️ No Stiiizy rows in {file_path.name}")
        return

    # Ensure numeric cost
    df["cost"] = pd.to_numeric(df["cost"], errors="coerce").fillna(0.0)

    # Kickback calculation (case-insensitive category matching)
    special_set = {c.lower() for c in SPECIAL_CATEGORIES}
    def calc_kickback(row):
        rate = 0.30 if str(row["category"]).strip() in SPECIAL_CATEGORIES else 0.15
        available = float(row.get("available", 1) or 1)
        return float(row["cost"]) * rate * available

    df["Kickback Owed"] = df.apply(calc_kickback, axis=1)
    def calc_totals(row):
        rate = 0.30 if str(row["category"]).strip() in SPECIAL_CATEGORIES else 0.15
        available = float(row.get("available", 1) or 1)
        unit_cost = float(row["cost"])
        total_cost = available * unit_cost
        kickback = total_cost * rate
        return pd.Series({"Total Cost": total_cost, "Kickback Owed": kickback})

    df[["Total Cost", "Kickback Owed"]] = df.apply(calc_totals, axis=1)
    # Keep only selected columns (in the requested order)
    keep = [c for c in KEEP_COLUMNS if (c in df.columns) or (c.lower() in df.columns)]
    # Align case of 'Kickback Owed' correctly if needed
    cols_map = {c.lower(): c for c in df.columns}
    normalized_keep = [cols_map.get(c.lower(), c) for c in keep]
    df = df[normalized_keep]

    # Prepare output path
    out_path = Path(OUTPUT_DIR) / file_path.with_suffix(".xlsx").name
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    # Write table so that header is exactly at HEADER_ROW
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Report", index=False, startrow=HEADER_ROW - 1)

    # Add summary + styling
    format_report(out_path, df)

    print(f"✅ Processed {file_path.name} → {out_path}")

# ───────────────────────── Styling / Summary ─────────────────────────
def format_report(xlsx_path: Path, df: pd.DataFrame):
    wb = load_workbook(xlsx_path)
    ws = wb["Report"]

    # Totals
# Totals
    total_cost = float(df["Total Cost"].sum())
    total_kick = float(df["Kickback Owed"].sum())
    item_count = int(df["available"].sum()) 

    # Summary block (rows 1–4)
    ws["A1"] = f"{TARGET_BRAND} Kickback Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"], ws["B2"] = "Total Cost", total_cost
    ws["A3"], ws["B3"] = "Total Kickback Owed", total_kick
    ws["A4"], ws["B4"] = "Total Items", item_count

    ws["B2"].number_format = '"$"#,##0.00'
    ws["B3"].number_format = '"$"#,##0.00'

    # Derive exact data range from DataFrame (avoid touching empty rows)
    ncols = df.shape[1]
    nrows = df.shape[0]
    header_row = HEADER_ROW
    first_data_row = header_row + 1
    last_data_row = header_row + nrows

    # Header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style only the actual header cells
    for col in range(1, ncols + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Money columns by header name
    header_names = [ws.cell(row=header_row, column=i).value for i in range(1, ncols + 1)]
    money_cols_idx = {i + 1 for i, name in enumerate(header_names)
                      if str(name).strip().lower() in {"cost", "kickback owed"}}

    # Style data rows (only the real data region)
    for row in range(first_data_row, last_data_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            if col in money_cols_idx:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Alternating row shading
            if (row - header_row) % 2 == 1:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Auto-fit widths based on header + data cells only (not the summary area)
    for col in range(1, ncols + 1):
        values = [ws.cell(row=header_row, column=col).value]
        for row in range(first_data_row, last_data_row + 1):
            values.append(ws.cell(row=row, column=col).value)
        text_lengths = [len(str(v)) for v in values if v is not None]
        width = max([12] + [t + 2 for t in text_lengths])  # minimum width 12
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze panes so ONLY the real header row is sticky
    ws.freeze_panes = ws[f"A{FREEZE_AT_ROW}"]

    wb.save(xlsx_path)

# ───────────────────────── Entrypoint ─────────────────────────
def main():
    Path(FILES_DIR).mkdir(exist_ok=True)
    Path(OUTPUT_DIR).mkdir(exist_ok=True)
    for csv_file in Path(FILES_DIR).glob("*.csv"):
        process_file(csv_file)

if __name__ == "__main__":
    main()
