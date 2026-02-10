#!/usr/bin/env python3
"""
AutoOrderPlanner.py
───────────────────
▪ Runs getCatalog.py → fresh CSVs in ./files/
▪ Runs getSalesReport.run_sales_report() for last <LOOKBACK> days
▪ Matches each “…_MV.csv” to salesMV.xlsx (same for LM / SV / LG)
▪ Builds order_plan_<brand>_<store>_<LOOKBACK>d.xlsx for every match
"""

from pathlib import Path
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox
import threading, time, subprocess, re
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import math
# ─── CONFIG ───────────────────────────────────────────────────────────────
LOOKBACK    = 30                        # days of sales history
FILES_DIR   = Path("files")             # both catalog CSV & sales XLSX
PLAN_DIR    = Path("order_plans")
MAX_AVAIL   = 2                         # >2 counted as available
WAIT_SALES  = 120                       # sec to wait for downloads
TARGET_STORES = {"MV"}    

CATALOG_COLS = ["Available", "Product", "Brand", "Category", "Cost"]
SALES_COLS   = {
    "product name": "Product",
    "inventory cost": "Cost",
    "total inventory sold": "QtySold",
    "order time": "Date",
}

CSV_TAG   = re.compile(r"_([A-Z]{2})\.csv$",  re.I)   # 2025-05-07_MV.csv
XLSX_TAG  = re.compile(r"sales([A-Z]{2})\.xlsx$", re.I) # salesMV.xlsx

# import Selenium runner
from getSalesReport import run_sales_report
# -------------------------------------------------------------------------
def need_refresh(folder: Path) -> bool:
    """Return True if FILES_DIR looks incomplete (odd # of files or empty)."""
    files = [p for p in folder.iterdir() if p.is_file()]
    return len(files) == 0 or len(files) % 2 == 1
def run_get_catalog():
    before = set(FILES_DIR.glob("*.csv"))
    subprocess.check_call(["python", "getCatalog.py", str(FILES_DIR)])
    t0 = time.time()
    while not (set(FILES_DIR.glob("*.csv")) - before):
        if time.time() - t0 > 60:
            raise TimeoutError("Catalog download timed out")
        time.sleep(2)

def parse_catalog() -> dict[tuple[str,str], pd.DataFrame]:
    """Return {(brand, store): df_of_available_rows}."""
    tables = {}
    for csv in FILES_DIR.glob("*.csv"):
        m = CSV_TAG.search(csv.name)
        if not m:
            continue
        store = m.group(1).upper()
        if store not in TARGET_STORES:         
            continue
        df = pd.read_csv(csv, usecols=lambda c: c in CATALOG_COLS)
        df = df[df["Available"] > MAX_AVAIL]
        df["Brand"] = df["Brand"].str.strip().str.lower()
        df["Store"] = store
        for brand, grp in df.groupby("Brand"):
            tables.setdefault((brand, store), []).append(grp)
    return {k: pd.concat(v, ignore_index=True) for k, v in tables.items()}

def load_sales(days: int) -> pd.DataFrame:
    end, start = datetime.today(), datetime.today()-timedelta(days=days)
    run_sales_report(start, end)
    t0 = time.time()
    frames = []
    while not frames:
        for x in FILES_DIR.glob("sales*.xlsx"):
            tag = XLSX_TAG.search(x.name)
            if not tag or tag.group(1).upper() not in TARGET_STORES:   # ← NEW guard
                continue
            if x.name.endswith(".crdownload"):
                continue
            frames.append(read_sales(x))
        if frames: break
        if time.time()-t0 > WAIT_SALES:
            raise TimeoutError("Sales download timed out")
        time.sleep(2)
    return pd.concat(frames, ignore_index=True)

def read_sales(path: Path) -> pd.DataFrame:
    """
    Robust loader for a Dutchie sales export:
      • accepts any upper/lower-case header variants
      • adds Store tag (MV / LM / SV / LG)
      • filters to the LOOKBACK window
    """
    df = pd.read_excel(path, header=4)           # read all columns

    # 1) normalise header labels
    df.columns = df.columns.str.strip().str.lower()

    # 2) friendly mapping → canonical names
    mapping = {
        "product name":        "Product",
        "name":                "Product",
        "inventory cost":      "Cost",
        "cost":                "Cost",
        "total inventory sold":"QtySold",
        "qty":                 "QtySold",
        "order time":          "Date",
        "order placed at":     "Date",
    }
    df = df.rename(columns=mapping)

    # 3) keep only the four columns we care about
    need = ["Product", "Cost", "QtySold", "Date"]
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise ValueError(f"{path.name}: missing expected cols {missing}")

    # 4) date parsing & time-window filter
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    cutoff = datetime.today() - timedelta(days=LOOKBACK)
    df = df[df["Date"] >= cutoff]

    # 5) tag the store
    m = XLSX_TAG.search(path.name)
    df["Store"] = m.group(1).upper() if m else "UNK"

    return df[need + ["Store"]]


def group_skus(inv: pd.DataFrame) -> pd.Series:
    """
    Assign the same gid to rows that share:
       • identical Cost
       • identical size token (1g / 14g / 100 mg …)
       • ≥70 % fuzzy-name similarity *after* the size token is removed
    """
    inv = inv.copy()
    inv["size_tok"] = inv["Product"].astype(str).apply(extract_size)

    inv_sorted = inv.sort_values(["Cost", "size_tok"]).reset_index(drop=True)
    idx2gid, gid = {}, 0

    for i, row in inv_sorted.iterrows():
        if i in idx2gid:
            continue
        idx2gid[i] = gid
        base_size = row.size_tok
        base_cost = row.Cost
        base_clean = row.Product.replace(base_size, "")

        for j in range(i + 1, len(inv_sorted)):
            r2 = inv_sorted.iloc[j]
            if r2.Cost != base_cost or r2.size_tok != base_size:
                break

            clean2 = r2.Product.replace(base_size, "")
            if fuzz.token_set_ratio(base_clean, clean2) >= 70:          # :contentReference[oaicite:1]{index=1}
                idx2gid[j] = gid
        gid += 1

    return pd.Series({inv_sorted.index[k]: v for k, v in idx2gid.items()},
                     name="gid")
def match_sales(sales, inv, gid_map):
    inv = inv.assign(gid=gid_map)
    buckets = {c: g[["Product","gid"]] for c,g in inv.groupby("Cost")}
    res=[]
    for _, row in sales.iterrows():
        b = buckets.get(row.Cost)
        best_gid,best= None,0
        if b is not None:
            for prod,gid in b.itertuples(False):
                sc=fuzz.token_set_ratio(prod,row.Product)
                if sc>best: best,best_gid=sc,gid
        res.append(best_gid if best>=70 else pd.NA)
    return pd.Series(res,name="gid")
# ────────────────────────────────────────────────────────────
# PATCH C – compute par & need with real quantities
# ────────────────────────────────────────────────────────────
WEIGHT_RE = re.compile(r'\b(\d+(?:\.\d+)?)(g|mg|ml)\b', re.I)
def extract_size(prod_name: str) -> str:
    """Return a normalised '14g', '1g', … token from a product string."""
    m = WEIGHT_RE.search(prod_name)
    return m.group(0).lower() if m else ""

def ceil_series(s: pd.Series) -> pd.Series:
    """Return a series where each numeric value is rounded *up* to an int."""
    return s.apply(lambda x: math.ceil(x) if pd.notna(x) else x)
def attach_sales(inv_df: pd.DataFrame, sales_df: pd.DataFrame) -> pd.DataFrame:
    """
    • Builds gid on the inventory side
    • Fuzzy-matches each sales row into those gids
    • Returns the inventory table with an extra 'UnitsSold' col
      (0 when that gid never appeared in the sales window)
    """
    # 1) make the gid map on the inventory
    gid_map = group_skus(inv_df)

    # 2) push gid into the sales slice
    sales_df = sales_df.copy()
    sales_df["gid"] = match_sales(sales_df, inv_df, gid_map)

    # 3) aggregate sales → UnitsSold per gid
    sold = (
        sales_df.dropna(subset=["gid"])
        .groupby("gid", as_index=False)["QtySold"]
        .sum()
        .rename(columns={"QtySold": "UnitsSold"})
    )

    # 4) bolt that back onto the inventory
    inv_with_gid = inv_df.assign(gid=gid_map)
    merged = inv_with_gid.merge(sold, on="gid", how="left").fillna({"UnitsSold": 0})

    return merged
def build_plan(inv_df: pd.DataFrame,
               sales_df: pd.DataFrame) -> pd.DataFrame:

    # inventory + total UnitsSold (from attach_sales)
    merged = attach_sales(inv_df, sales_df)

    # ---------------------------------------------------------
    # 1) work out the span of sales for every gid
    # ---------------------------------------------------------
    gmap = group_skus(inv_df)
    sales_df = sales_df.copy()
    sales_df["gid"] = match_sales(sales_df, inv_df, gmap)

    span = (sales_df.dropna(subset=["gid"])
                     .groupby("gid")
                     .agg(first_date=("Date", "min"),
                          last_date=("Date",  "max"))
           )
    span["days_in_stock"] = (span["last_date"] -
                             span["first_date"]).dt.days + 1

    # SKUs that never sold in the look-back window
    span["days_in_stock"].fillna(LOOKBACK, inplace=True)

    # ---------------------------------------------------------
    # 2) merge just the days_in_stock field back
    # ---------------------------------------------------------
    merged = (merged.merge(span[["days_in_stock"]],
                           on="gid", how="left")
                    .fillna({"days_in_stock": LOOKBACK}))

    # ---------------------------------------------------------
    # 3) aggregate to one row per gid
    # ---------------------------------------------------------
    grp = (merged.groupby("gid", as_index=False)
                  .agg(ExampleSKU=("Product", "first"),
                       Available=("Available", "sum"),
                       UnitsSold=("UnitsSold", "sum"),
                       days_in_stock=("days_in_stock", "first")))

    # ---------------------------------------------------------
    # 4) par levels – round-UP to whole units
    # ---------------------------------------------------------
    grp["avg_daily"] = ceil_series(grp["UnitsSold"] / grp["days_in_stock"])

    for w in (7, 14):
        grp[f"par_{w}d"]  = grp["avg_daily"] * w
        grp[f"need_{w}d"] = (grp[f"par_{w}d"] - grp["Available"]).clip(lower=0)

    return grp.sort_values("need_14d", ascending=False)

def prettify(xlsx: Path):
    wb=load_workbook(xlsx); ws=wb.active
    fill=PatternFill("solid",start_color="D9D9D9",end_color="D9D9D9")
    for c in ws[1]: c.font=Font(bold=True); c.fill=fill
    ws.freeze_panes="A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(
            len(str(c.value)) if c.value else 0 for c in col)+2
    wb.save(xlsx)

def run_pipeline():
    try:
        # ── decide whether to refresh files
        refresh = need_refresh(FILES_DIR)
        if refresh:
            print("🔄  Refreshing catalog + sales …")
            # wipe folder
            for p in FILES_DIR.iterdir():
                if p.is_file(): p.unlink()
            run_get_catalog()                      # fresh CSVs
            all_sales = load_sales(LOOKBACK)       # fresh sales (also downloads)
        else:
            print("✅  Even # of files detected – using existing data.")
            all_sales = pd.concat(
                read_sales(x) for x in FILES_DIR.glob("sales*.xlsx")
                if XLSX_TAG.search(x.name) and
                XLSX_TAG.search(x.name).group(1).upper() in TARGET_STORES   # ← guard
        )
        
        inv_tables = parse_catalog()               # always read CSVs
        PLAN_DIR.mkdir(exist_ok=True)

        for (brand, store), inv_df in inv_tables.items():
            brand_sales = all_sales[(all_sales.Store == store) &
                                    all_sales.Product.str.contains(brand,
                                                                   case=False,
                                                                   na=False)]
            if brand_sales.empty:
                continue
            plan = build_plan(inv_df, brand_sales)
            out  = PLAN_DIR / f"{brand}_{store}_{LOOKBACK}d.xlsx"
            plan.to_excel(out, index=False)
            prettify(out)
            print("✓", out.name)

        messagebox.showinfo("Done", f"Plans saved in {PLAN_DIR.resolve()}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def main():
    root=tk.Tk(); root.title("Auto Order Planner")
    tk.Label(root,text="One click → catalog & sales → order plans",
             font=("Arial",11)).pack(pady=20)
    tk.Button(root,text="Generate Order Plans",
              font=("Arial",12,"bold"),bg="lightgreen",
              command=lambda:threading.Thread(target=run_pipeline,daemon=True).start()
             ).pack(pady=10,ipadx=12,ipady=5)
    root.mainloop()

if __name__=="__main__":
    main()
