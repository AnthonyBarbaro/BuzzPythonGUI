#!/usr/bin/env python3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────── Config ─────────────────────────
FILES_DIR = "files"
OUTPUT_DIR = "doneReports"

TARGET_BRAND = "Kushy Punch"
TARGET_CATEGORY = "Edibles"

# Columns to show in the final report (order matters)
KEEP_COLUMNS = [
    "available",                  # current on-hand units
    "product",
    "category",
    "price",                      # unit retail price
    "cost",                       # current unit cost
    "Target Unit Cost",           # price / 4 (for 50% margin on BOGO pairs)
    "Delta to Target",            # cost - target (positive = too expensive)
    "Delta Cash (Avail×Δ)",       # max(Delta,0) * available
    "Current Margin % (BOGO)"     # (price - 2*cost) / price
]

# Layout: summary (rows 1–4), one blank spacer (row 5), table header (row 6)
SUMMARY_ROWS_USED = 4
SPACER_ROWS = 1
HEADER_ROW = SUMMARY_ROWS_USED + SPACER_ROWS + 1  # -> 6 (header at row 6)
FREEZE_AT_ROW = HEADER_ROW + 1                    # -> 7 (first data row)

# ───────────────────────── Helpers ─────────────────────────
def _to_numeric(series: pd.Series) -> pd.Series:
    """Convert money-like strings ($, commas, spaces) to numeric safely."""
    return (
        series.astype(str)
              .str.replace(r"[^0-9.\-]", "", regex=True)
              .replace({"": None, ".": None})
              .astype(float)
    )

def _first_existing(df: pd.DataFrame, candidates) -> str:
    """Return the first column name from candidates that exists in df (case-insensitive)."""
    cols_map = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols_map:
            return cols_map[c.lower()]
    return ""

def _money_fmt(cell):
    cell.number_format = '"$"#,##0.00'

def _pct_fmt(cell):
    cell.number_format = "0.00%"

# ───────────────────────── Core ─────────────────────────
def process_file(file_path: Path) -> dict:
    """
    Read a CSV, filter Kushy Punch Edibles, compute BOGO targets, write Excel.
    Returns a dict with per-file totals used to build the overall summary.
    """
    df = pd.read_csv(file_path)

    # Normalize column names (only names; leave values as-is)
    df.columns = df.columns.str.strip().str.lower()

    # Required columns: brand, category, cost, available, and price (or location price)
    brand_col    = _first_existing(df, ["brand"])
    category_col = _first_existing(df, ["category"])
    cost_col     = _first_existing(df, ["cost"])
    avail_col    = _first_existing(df, ["available"])
    price_col    = _first_existing(df, ["price"]) or _first_existing(df, ["location price", "location_price"])
    product_col  = _first_existing(df, ["product"])

    if not all([brand_col, category_col, cost_col, avail_col, price_col, product_col]):
        print(f"⚠️ {file_path.name} is missing required columns. Skipping.")
        return {"file": file_path.name, "credit_needed": 0.0, "skus": 0}

    # Brand / Category filters (case-insensitive)
    df = df[
        (df[brand_col].astype(str).str.strip().str.casefold() == TARGET_BRAND.casefold()) &
        (df[category_col].astype(str).str.strip().str.casefold() == TARGET_CATEGORY.casefold())
    ].copy()

    if df.empty:
        print(f"ℹ️ No rows for {TARGET_BRAND} {TARGET_CATEGORY} in {file_path.name}")
        return {"file": file_path.name, "credit_needed": 0.0, "skus": 0}

    # Clean numerics
    df["__price"]     = _to_numeric(df[price_col])
    df["__cost"]      = _to_numeric(df[cost_col])
    df["__available"] = pd.to_numeric(df[avail_col], errors="coerce").fillna(0).astype(float)

    # Business filters
    df = df[df["__available"] > 0]                    # in stock
    df = df[~(df["__cost"].round(2) == 0.01)]         # remove penny cost
    df = df[df["__price"] > 0]                        # must have a positive price
    if df.empty:
        print(f"ℹ️ After stock/cost filters, nothing left for {file_path.name}")
        return {"file": file_path.name, "credit_needed": 0.0, "skus": 0}

    # Compute targets for BOGO 50% margin:
    # Target Unit Cost = price / 4  (pair costs price/2; margin = 50%)
    df["Target Unit Cost"] = (df["__price"] / 4.0).round(2)

    # Delta to Target (positive = over target)
    df["Delta to Target"] = (df["__cost"] - df["Target Unit Cost"]).round(2)

    # Delta Cash = max(Delta, 0) × available   (how much extra you're paying in total)
    df["Delta Cash (Avail×Δ)"] = (df["Delta to Target"].clip(lower=0) * df["__available"]).round(2)

    # Current Margin % under BOGO = (price - 2*cost) / price
    df["Current Margin % (BOGO)"] = ((df["__price"] - 2.0 * df["__cost"]) / df["__price"]).clip(lower=-10, upper=10)

    # Build output table
    output = pd.DataFrame({
        "available": df["__available"].astype(int),
        "product": df[product_col],
        "category": df[category_col],
        "price": df["__price"].round(2),
        "cost": df["__cost"].round(2),
        "Target Unit Cost": df["Target Unit Cost"],
        "Delta to Target": df["Delta to Target"],
        "Delta Cash (Avail×Δ)": df["Delta Cash (Avail×Δ)"],
        "Current Margin % (BOGO)": df["Current Margin % (BOGO)"]
    })

    # Keep ordered subset
    keep = [c for c in KEEP_COLUMNS if c in output.columns]
    output = output[keep]

    # Prepare output path
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    out_path = Path(OUTPUT_DIR) / f"KushyPunch_BOGO_{file_path.stem}.xlsx"

    # Write so header is exactly at HEADER_ROW (row 6)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="Report", index=False, startrow=HEADER_ROW - 1)

    # Format + summary
    credit_needed = format_report(out_path, output)
    print(f"✅ Processed {file_path.name} → {out_path}  |  Credit Needed: ${credit_needed:,.2f}")

    return {
        "file": file_path.name,
        "credit_needed": float(credit_needed),
        "skus": int(output.shape[0])
    }

# ───────────────────────── Styling / Summary ─────────────────────────
def format_report(xlsx_path: Path, df: pd.DataFrame) -> float:
    wb = load_workbook(xlsx_path)
    ws = wb["Report"]

    # Summary metric: sum of positive Delta Cash
    total_credit = float(df.loc[df["Delta Cash (Avail×Δ)"] > 0, "Delta Cash (Avail×Δ)"].sum())

    # Title & single summary line
    ws["A1"] = f"{TARGET_BRAND} BOGO Cost Targets ({TARGET_CATEGORY})"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"], ws["B2"] = "Total Credit Needed to Hit Target", total_credit
    _money_fmt(ws["B2"])

    # (Rows A3/A4 intentionally left blank so the table header stays at row 6)

    # Table range
    ncols = df.shape[1]
    nrows = df.shape[0]
    header_row = HEADER_ROW
    first_data_row = header_row + 1
    last_data_row = header_row + nrows

    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col in range(1, ncols + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Which columns are money / percent?
    headers = [ws.cell(row=header_row, column=i).value for i in range(1, ncols + 1)]
    money_headers = {
        "price", "cost", "target unit cost",
        "delta to target", "delta cash (avail×Δ)", "delta cash (availxΔ)"
    }
    pct_headers = {"current margin % (bogo)"}

    money_cols = {i + 1 for i, h in enumerate(headers) if str(h).strip().lower() in money_headers}
    pct_cols   = {i + 1 for i, h in enumerate(headers) if str(h).strip().lower() in pct_headers}

    # Data styling
    for row in range(first_data_row, last_data_row + 1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            header_text = str(ws.cell(row=header_row, column=col).value).strip().lower()
            if col in money_cols:
                _money_fmt(cell)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in pct_cols:
                _pct_fmt(cell)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                align = "right" if header_text == "available" else "left"
                cell.alignment = Alignment(horizontal=align, vertical="center")

            # Alternating row shading
            if (row - header_row) % 2 == 1:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Auto-fit widths
    for col in range(1, ncols + 1):
        values = [ws.cell(row=header_row, column=col).value]
        for row in range(first_data_row, last_data_row + 1):
            values.append(ws.cell(row=row, column=col).value)
        texts = [str(v) for v in values if v is not None]
        width = max([12] + [len(t) + 2 for t in texts])
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze so ONLY the table header (row 6) sticks
    ws.freeze_panes = ws[f"A{FREEZE_AT_ROW}"]

    wb.save(xlsx_path)
    return total_credit

# ───────────────────────── Entrypoint ─────────────────────────
def main():
    Path(FILES_DIR).mkdir(exist_ok=True)
    Path(OUTPUT_DIR).mkdir(exist_ok=True)

    per_file = []
    for csv_file in Path(FILES_DIR).glob("*.csv"):
        per_file.append(process_file(csv_file))

    # Build overall summary across all processed files
    df_summary = pd.DataFrame(per_file)
    df_summary = df_summary[df_summary["skus"] > 0] if not df_summary.empty else df_summary

    if df_summary is None or df_summary.empty:
        print("ℹ️ No matching data found in any CSVs under 'files/'.")
        return

    # Save the summary CSV and print the grand total
    summary_path = Path(OUTPUT_DIR) / "KushyPunch_BOGO_SUMMARY.csv"
    df_summary.to_csv(summary_path, index=False)

    grand_total = float(df_summary["credit_needed"].sum())
    print("\n──────── SUMMARY ────────")
    for _, r in df_summary.iterrows():
        print(f"{r['file']}: ${r['credit_needed']:,.2f}")
    print(f"Grand Total Credit Needed: ${grand_total:,.2f}")
    print(f"Saved: {summary_path}")

if __name__ == "__main__":
    main()
